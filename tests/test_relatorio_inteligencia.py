"""
Testes do RELATÓRIO DE INTELIGÊNCIA de fornecedor (compliance_agent.reporting.inteligencia)
e do mapa canônico de UGs (compliance_agent.ugs).

Rodam contra a base local `data/compliance.db` (sem rede): o enriquecimento por APIs públicas é
desligado via timeout curto, então validamos a espinha REAL (OBs/contratos/UGs) de forma determinística.

Como rodar:
    cd ~/JFN && .venv/bin/python -m pytest tests/test_relatorio_inteligencia.py -v
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# enriquecimento (rede) desligado: validamos só o que é determinístico (compliance.db)
os.environ.setdefault("JFN_RELATORIO_ENRIQUECE_TIMEOUT", "0.01")

from compliance_agent.reporting import inteligencia as I  # noqa: E402
from compliance_agent import ugs  # noqa: E402

CNPJ_MGS = "19088605000104"
_DB = _ROOT / "data" / "compliance.db"
_tem_db = _DB.exists()
skip_sem_db = pytest.mark.skipif(not _tem_db, reason="data/compliance.db ausente neste ambiente")


# ───────────────────────────── helpers de formatação ─────────────────────────────

def test_so_digitos_e_fmt_cnpj():
    assert I.so_digitos("19.088.605/0001-04") == CNPJ_MGS
    assert I.fmt_cnpj(CNPJ_MGS) == "19.088.605/0001-04"
    assert I.fmt_cnpj("123") == "123"  # inválido devolve como veio


def test_moeda_padrao_br():
    assert I.moeda(1234567.89) == "1.234.567,89"
    assert I.moeda(0) == "0,00"
    assert I.moeda(None) == "0,00"


# ───────────────────────────── UGs canônicas (aprendizado ITERJ) ─────────────────────────────

def test_ug_iterj_e_canonica():
    # UG 133100 deve ser reconhecida como ITERJ, não como a Secretaria de Infraestrutura
    nome = ugs.nome_canonico("133100")
    assert "ITERJ" in nome
    rot = ugs.rotulo("133100", "Secretaria de Estado de Infraestrutura e Obras")
    assert "ITERJ" in rot and "133100" in rot
    # alias entre sistemas de numeração documentado
    assert ugs.ALIASES["133100"]["siafe_rio2"] == "270042"


def test_ug_titulo_conectivos_minusculos():
    t = ugs._titulo("FUND.ESP.DO CORPO DE BOMBEIROS")
    assert " de " in f" {t} " or "de" in t.lower()
    assert "DE BOMBEIROS" not in t  # conectivo não fica em caixa alta


# ───────────────────────────── resolução por nome (parcial) ─────────────────────────────

@skip_sem_db
def test_resolver_por_nome_parcial_mgs():
    cands = I.buscar_candidatos("MGS")
    assert cands, "deveria achar ao menos a MGS"
    assert any(c["cnpj"] == CNPJ_MGS for c in cands)


@skip_sem_db
def test_resolver_por_cnpj_direto():
    cands = I.buscar_candidatos("19.088.605/0001-04")
    assert len(cands) == 1 and cands[0]["cnpj"] == CNPJ_MGS


@skip_sem_db
def test_nome_ambiguo_lista_candidatos():
    # "clean" casa com várias empresas -> mais de um candidato (dispara desambiguação no montar)
    cands = I.buscar_candidatos("clean")
    assert len(cands) >= 2


# ───────────────────────────── pagamentos por ano (requisito Jorge) ─────────────────────────────

@skip_sem_db
def test_pagamentos_por_ano_mgs():
    p = I.consultar_pagamentos(CNPJ_MGS)
    assert p["tem_dados"]
    assert p["anos"] == sorted(p["anos"])
    assert set(p["anos"]) >= {2023, 2024, 2025, 2026}
    # cada ano tem suas linhas individuais (podem ser >12)
    for ano in p["anos"]:
        bloco = p["por_ano"][ano]
        assert bloco["n"] == len(bloco["linhas"])
        assert bloco["n"] >= 1
        # total do ano = soma das linhas
        soma = round(sum(l["valor"] for l in bloco["linhas"]), 2)
        assert abs(soma - round(bloco["total"], 2)) < 0.01
    # total geral coerente
    assert abs(p["total_geral"] - sum(p["por_ano"][a]["total"] for a in p["anos"])) < 0.01


@skip_sem_db
def test_iterj_separado_da_secretaria_na_concentracao():
    # após a correção de UG, ITERJ aparece como órgão próprio (não diluído na Secretaria)
    p = I.consultar_pagamentos(CNPJ_MGS)
    orgaos = " | ".join(p["por_orgao_geral"].keys())
    assert "ITERJ" in orgaos


@skip_sem_db
def test_hhi_calculo():
    p = I.consultar_pagamentos(CNPJ_MGS)
    hhi = p["hhi"]
    assert hhi["nivel"] in ("BAIXA", "MODERADA", "ALTA")
    assert 0 <= hhi["top_share"] <= 100


# ───────────────────────────── geração ponta a ponta ─────────────────────────────

@skip_sem_db
def test_montar_gera_md_e_pdf(tmp_path, monkeypatch):
    import asyncio
    res = asyncio.run(I.montar(cnpj=CNPJ_MGS))
    assert res["ok"] is True
    assert res["cnpj"] == CNPJ_MGS
    assert res["fonte"] == "REAL"
    # markdown salvo e com a seção de tabelas por ano
    md_path = Path(res["path_md"])
    assert md_path.exists()
    texto = md_path.read_text(encoding="utf-8")
    assert "PAGAMENTOS (ORDENS BANCÁRIAS) POR ANO" in texto
    assert "### Exercício 2025" in texto
    assert "ITERJ" in texto  # UG corrigida aparece
    # PDF gerado e não-vazio
    if res["path_pdf"]:
        assert Path(res["path_pdf"]).stat().st_size > 1000


@skip_sem_db
def test_montar_cnpj_invalido():
    import asyncio
    res = asyncio.run(I.montar(cnpj="000"))
    assert res["ok"] is False


# ───────────────────────────── relatório de ÓRGÃO ─────────────────────────────

from compliance_agent.reporting import inteligencia_orgao as IO  # noqa: E402


@skip_sem_db
def test_orgao_resolve_iterj_por_nome():
    cands = IO.buscar_orgaos("iterj")
    assert cands and any(c["ug"] == "133100" for c in cands)


@skip_sem_db
def test_orgao_iterj_pagamentos_por_ano():
    p = IO.consultar_orgao("133100")
    assert p["tem_dados"]
    assert set(p["anos"]) >= {2023, 2024, 2025, 2026}
    assert p["n_fornecedores"] >= 1
    # concentração por favorecido coerente
    soma = sum(p["por_favorecido_geral"].values())
    assert abs(soma - p["total_geral"]) < 1.0


@skip_sem_db
def test_orgao_montar_gera_md_pdf_com_parecer():
    res = IO.montar(orgao="iterj")
    assert res["ok"] is True and res["ug"] == "133100"
    md = Path(res["path_md"]).read_text(encoding="utf-8")
    assert "RELATÓRIO DE INTELIGÊNCIA — ITERJ" in md
    assert "PARECER PRELIMINAR" in md  # branding ac5f451: sem "DO JFN" nos entregáveis
    assert "PAGAMENTOS (ORDENS BANCÁRIAS) POR ANO" in md
    if res["path_pdf"]:
        assert Path(res["path_pdf"]).stat().st_size > 1000


@skip_sem_db
def test_planilha_excel_gerada(tmp_path):
    from compliance_agent.reporting import planilha
    import openpyxl
    p = I.consultar_pagamentos(CNPJ_MGS)
    ctx = {"nome": "MGS", "cnpj_fmt": I.fmt_cnpj(CNPJ_MGS), "data": "2026-06-06", "pagamentos": p}
    dest = str(tmp_path / "t.xlsx")
    planilha.gerar(ctx, dest, modo="fornecedor")
    wb = openpyxl.load_workbook(dest)
    assert "Pagamentos (OBs)" in wb.sheetnames
    ws = wb["Pagamentos (OBs)"]
    assert "Pagamentos" in ws.tables          # é uma Tabela do Excel (interativa)
    assert ws.freeze_panes == "A2"            # cabeçalho congelado
    assert ws.max_row - 1 == p["n_geral"]     # uma linha por OB


def test_parecer_fornecedor_sem_dados_nao_quebra():
    ctx = {"nome": "X", "score": 0, "risco": "—",
           "pagamentos": {"tem_dados": False}}
    txt = I.parecer_fornecedor(ctx)
    assert "Sem Ordens Bancárias" in txt
