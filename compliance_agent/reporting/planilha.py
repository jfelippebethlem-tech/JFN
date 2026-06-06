# -*- coding: utf-8 -*-
"""
PLANILHA EXCEL INTERATIVA dos pagamentos (OBs) — anexo padrão dos relatórios do JFN.

Boas práticas de planilha de auditoria aplicadas (referências: guias de data-cleaning/analytics do
ACFE/IIA e do próprio Excel):
  - **Tabela do Excel** (ListObject) com **autofiltro** e linhas zebradas → o usuário filtra/ordena por
    qualquer campo (ano, órgão, fornecedor, valor) na hora.
  - **Painel congelado** no cabeçalho (freeze panes) para rolar mantendo os títulos.
  - **Formato de moeda** (R$ #.##0,00) e **data** corretos (valores numéricos reais, não texto) → permite
    somar, filtrar por faixa e construir tabela dinâmica.
  - **Aba Resumo** com KPIs + agregação por ano e por órgão/fornecedor (pronta para gráfico/pivot).
  - **Formatação condicional**: estornos (R$ 0,00) destacados; **barras de dados** na coluna de valor.
  - Uma linha por OB; nomes de coluna claros e únicos (regra de ouro: dado tabular limpo = pivotável).

USO:
    from compliance_agent.reporting import planilha
    planilha.gerar(ctx, "reports/arquivo.xlsx", modo="fornecedor")   # ou modo="orgao"
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

# paleta JFN
_AZUL = "1F2E4D"
_AZUL_CLARO = "2D3C5A"
_CINZA = "F4F6FA"
_VERMELHO = "FFC7CE"
_FMT_MOEDA = 'R$ #,##0.00'


def _data(s):
    try:
        return date.fromisoformat((s or "")[:10])
    except Exception:
        return s or ""


def gerar(ctx: dict, destino: str, modo: str = "fornecedor") -> str:
    """Gera a planilha interativa. modo='fornecedor' (coluna Órgão) ou 'orgao' (coluna Fornecedor)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    from openpyxl.utils import get_column_letter

    p = ctx["pagamentos"]
    titulo = ctx.get("nome", "")
    subt = ctx.get("cnpj_fmt") or (f"UG {ctx.get('ug','')}" if ctx.get("ug") else "")

    wb = Workbook()

    # estilos reutilizáveis
    f_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    f_sub = Font(name="Calibri", size=10, color="FFFFFF")
    f_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fill_az = PatternFill("solid", fgColor=_AZUL)
    fill_az2 = PatternFill("solid", fgColor=_AZUL_CLARO)
    fill_cinza = PatternFill("solid", fgColor=_CINZA)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="D9D9D9")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Relatório de Inteligência — {titulo}"
    ws["A1"].font = f_titulo; ws["A1"].fill = fill_az; ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:E2")
    ws["A2"] = f"{subt}  ·  Pagamentos (Ordens Bancárias) · gerado em {ctx.get('data','')}"
    ws["A2"].font = f_sub; ws["A2"].fill = fill_az2
    ws.row_dimensions[2].height = 16

    linha = 4
    if p.get("tem_dados"):
        # KPIs
        kpis = [("Total pago (R$)", p["total_geral"], _FMT_MOEDA),
                ("Nº de OBs", p["n_geral"], "#,##0"),
                (("Fornecedores" if modo == "orgao" else "Órgãos"),
                 p.get("n_fornecedores") if modo == "orgao" else len(p.get("por_orgao_geral", {})), "#,##0"),
                ("HHI (concentração)", p["hhi"].get("indice"), "#,##0.0"),
                ("Maior fatia (%)", p["hhi"].get("top_share"), "#,##0.0")]
        ws[f"A{linha}"] = "Indicadores"; ws[f"A{linha}"].font = Font(bold=True, size=11, color=_AZUL)
        linha += 1
        for nome, val, fmt in kpis:
            ws[f"A{linha}"] = nome; ws[f"A{linha}"].font = Font(bold=True)
            c = ws[f"C{linha}"]; c.value = val; c.number_format = fmt; c.alignment = right
            linha += 1
        linha += 1

        # por ano
        ws[f"A{linha}"] = "Pagamentos por exercício"; ws[f"A{linha}"].font = Font(bold=True, size=11, color=_AZUL)
        linha += 1
        cab_ano = ["Exercício", "Nº de OBs"] + (["Fornecedores"] if modo == "orgao" else []) + ["Valor pago (R$)"]
        for j, h in enumerate(cab_ano):
            c = ws.cell(row=linha, column=1 + j, value=h); c.font = f_hdr; c.fill = fill_az; c.alignment = center; c.border = borda
        linha += 1
        ini_ano = linha
        for a in p["anos"]:
            b = p["por_ano"][a]
            row = [a, b["n"]]
            if modo == "orgao":
                row.append(len({l.get("cnpj") for l in b["linhas"]}))
            row.append(b["total"])
            for j, v in enumerate(row):
                c = ws.cell(row=linha, column=1 + j, value=v); c.border = borda
                if j == len(row) - 1:
                    c.number_format = _FMT_MOEDA; c.alignment = right
                elif j == 0:
                    c.alignment = center
            linha += 1
        # barra de dados na coluna de valor da tabela por ano
        col_val = get_column_letter(len(cab_ano))
        ws.conditional_formatting.add(f"{col_val}{ini_ano}:{col_val}{linha-1}",
                                      DataBarRule(start_type="min", end_type="max", color="4F81BD"))

    # larguras da aba resumo
    for col, w in zip("ABCDE", (26, 16, 18, 16, 16)):
        ws.column_dimensions[col].width = w

    # ── Aba 2: Pagamentos (OBs) — tabela interativa ───────────────────────────
    ws2 = wb.create_sheet("Pagamentos (OBs)")
    ws2.sheet_view.showGridLines = False
    if modo == "orgao":
        cols = ["Ano", "Nº OB", "Data pagamento", "Fornecedor", "CNPJ", "Valor (R$)"]
    else:
        cols = ["Ano", "Nº OB", "Data pagamento", "Órgão (UG)", "Valor (R$)"]
    for j, h in enumerate(cols):
        c = ws2.cell(row=1, column=1 + j, value=h)
        c.font = f_hdr; c.fill = fill_az; c.alignment = center
    ridx = 2
    if p.get("tem_dados"):
        for a in p["anos"]:
            for ln in p["por_ano"][a]["linhas"]:
                if modo == "orgao":
                    vals = [a, ln["numero_ob"], _data(ln["data"]), ln["favorecido"],
                            ("'" + ln["cnpj"]) if ln.get("cnpj") else "", ln["valor"]]
                else:
                    vals = [a, ln["numero_ob"], _data(ln["data"]), ln["orgao"], ln["valor"]]
                for j, v in enumerate(vals):
                    c = ws2.cell(row=ridx, column=1 + j, value=v)
                    if cols[j] == "Valor (R$)":
                        c.number_format = _FMT_MOEDA; c.alignment = right
                    elif cols[j] == "Data pagamento":
                        c.number_format = "dd/mm/yyyy"; c.alignment = center
                    elif cols[j] == "Ano":
                        c.alignment = center
                ridx += 1
    ultima = ridx - 1
    ncol = len(cols)
    col_letter = get_column_letter(ncol)

    if ultima >= 2:
        # Tabela do Excel (autofiltro + zebra) — o "interativo"
        ref = f"A1:{col_letter}{ultima}"
        tab = Table(displayName="Pagamentos", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws2.add_table(tab)
        ws2.freeze_panes = "A2"  # congela cabeçalho
        # estorno (valor 0) destacado em vermelho
        ws2.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ultima}",
            CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=_VERMELHO)))
        # barra de dados na coluna de valor
        ws2.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{ultima}",
            DataBarRule(start_type="min", end_type="max", color="63A0D8"))

    # larguras
    larg = {"Ano": 8, "Nº OB": 16, "Data pagamento": 16, "Órgão (UG)": 52, "Fornecedor": 46, "CNPJ": 20, "Valor (R$)": 18}
    for j, h in enumerate(cols):
        ws2.column_dimensions[get_column_letter(1 + j)].width = larg.get(h, 18)

    # ── Aba 3: Concentração ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Concentração")
    ws3.sheet_view.showGridLines = False
    rotulo = "Fornecedor" if modo == "orgao" else "Órgão (UG)"
    conc = p.get("por_favorecido_geral" if modo == "orgao" else "por_orgao_geral", {}) if p.get("tem_dados") else {}
    cab3 = [rotulo, "Valor (R$)", "% do total"]
    for j, h in enumerate(cab3):
        c = ws3.cell(row=1, column=1 + j, value=h); c.font = f_hdr; c.fill = fill_az; c.alignment = center
    tot = (p.get("total_geral") or 1) if p.get("tem_dados") else 1
    r = 2
    for nome, val in conc.items():
        ws3.cell(row=r, column=1, value=nome)
        cv = ws3.cell(row=r, column=2, value=val); cv.number_format = _FMT_MOEDA; cv.alignment = right
        cp = ws3.cell(row=r, column=3, value=val / tot); cp.number_format = "0.0%"; cp.alignment = right
        r += 1
    if r > 2:
        tab3 = Table(displayName="Concentracao", ref=f"A1:C{r-1}")
        tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws3.add_table(tab3)
        ws3.freeze_panes = "A2"
        ws3.conditional_formatting.add(f"B2:B{r-1}", DataBarRule(start_type="min", end_type="max", color="63A0D8"))
    ws3.column_dimensions["A"].width = 52; ws3.column_dimensions["B"].width = 18; ws3.column_dimensions["C"].width = 12

    # ── Aba 4 (só órgão): Por Fornecedor — OBs AGRUPADAS por fornecedor (já organizado) ──
    if modo == "orgao" and p.get("tem_dados"):
        ws4 = wb.create_sheet("Por Fornecedor")
        ws4.sheet_view.showGridLines = False
        ws4.sheet_properties.outlinePr.summaryBelow = False  # subtotal do fornecedor ACIMA das suas OBs
        porf: dict = defaultdict(list)
        for a in p["anos"]:
            for ln in p["por_ano"][a]["linhas"]:
                porf[ln["favorecido"]].append({**ln, "ano": a})
        ordem = sorted(porf, key=lambda fav: sum(x["valor"] for x in porf[fav]), reverse=True)
        cab4 = ["Fornecedor / Nº OB", "CNPJ", "Ano", "Data", "Valor (R$)"]
        for j, h in enumerate(cab4):
            c = ws4.cell(row=1, column=1 + j, value=h); c.font = f_hdr; c.fill = fill_az; c.alignment = center
        ws4.freeze_panes = "A2"
        r = 2
        bold = Font(bold=True)
        fill_grp = PatternFill("solid", fgColor="E8ECF5")
        for fav in ordem:
            obs = sorted(porf[fav], key=lambda x: (x["ano"], str(x["data"])))
            tot_f = sum(x["valor"] for x in obs)
            cnpj = next((x["cnpj"] for x in obs if x.get("cnpj")), "")
            cf = ws4.cell(row=r, column=1, value=fav); cf.font = bold; cf.fill = fill_grp
            cc = ws4.cell(row=r, column=2, value=("'" + cnpj) if cnpj else ""); cc.fill = fill_grp
            cn = ws4.cell(row=r, column=4, value=f"{len(obs)} OBs"); cn.alignment = right; cn.fill = fill_grp; cn.font = bold
            cv = ws4.cell(row=r, column=5, value=tot_f); cv.number_format = _FMT_MOEDA; cv.alignment = right; cv.font = bold; cv.fill = fill_grp
            ws4.cell(row=r, column=3).fill = fill_grp
            r += 1
            for x in obs:
                ws4.cell(row=r, column=1, value="    " + str(x["numero_ob"]))
                ws4.cell(row=r, column=3, value=x["ano"]).alignment = center
                cd = ws4.cell(row=r, column=4, value=_data(x["data"])); cd.number_format = "dd/mm/yyyy"; cd.alignment = center
                cvv = ws4.cell(row=r, column=5, value=x["valor"]); cvv.number_format = _FMT_MOEDA; cvv.alignment = right
                ws4.row_dimensions[r].outline_level = 1  # agrupável: clique no "-" para colapsar o fornecedor
                r += 1
        for col, w in zip("ABCDE", (50, 20, 8, 14, 18)):
            ws4.column_dimensions[col].width = w

    Path(destino).parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino
