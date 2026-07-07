# -*- coding: utf-8 -*-
"""Relatório Câmara × Prefeitura do Rio — indícios de duplo vínculo por nome + mapa de gabinetes.

Produto da casa (Kroll): md + pdf via ``reporting/render_html``. Honestidade dura:
sem CPF, casamento por nome é INDÍCIO (homônimo possível), jamais acusação. Vínculo
efetivo/carreira no Executivo (Guarda Municipal, Professor, concursado) sobreposto a
posto comissionado na Câmara é o indício mais forte de acumulação vedada (CF art. 37,
XVI/XVII) — a apurar por CPF/RH, nunca afirmado aqui.
"""
from __future__ import annotations

import logging
import html as _html
from datetime import datetime, timezone

from compliance_agent.pcrj import db as _db

# Palavras que sugerem vínculo EFETIVO/carreira no Executivo (↑ concern de acúmulo).
_EFETIVO = ("GUARDA MUNICIPAL", "PROFESSOR", "MEDICO", "MÉDICO", "ENFERMEIR",
            "AGENTE DE ADMINISTRA", "AUXILIAR DE ENFERM", "FISCAL", "ANALISTA",
            "TECNICO", "TÉCNICO", "ASSISTENTE SOCIAL", "PROCURADOR", "AUDITOR")


logger = logging.getLogger(__name__)


def _e(x) -> str:
    return _html.escape(str(x or ""))


def _vinculo_efetivo(cargo_pcrj: str) -> bool:
    c = (cargo_pcrj or "").upper()
    return any(k in c for k in _EFETIVO)


def _cessao_provavel(v: dict) -> bool:
    """Cessão/requisição (vínculo ÚNICO, NÃO acúmulo): o servidor da Prefeitura está
    formalmente 'à disposição' da Câmara, OU a Câmara o lista como 'Requisitado'. Nesses
    casos as duas folhas retratam o MESMO posto — não é duplo vínculo remunerado."""
    orgao = (v.get("orgao_pcrj") or "").upper()
    vinc = (v.get("vinculo_camara") or "").upper()
    return ("DISP" in orgao or "DISPOSI" in orgao or "CMRJ" in orgao
            or "REQUISITAD" in vinc)


def _coletar(con) -> dict:
    tot_camara = con.execute(
        "SELECT COUNT(DISTINCT nome_norm) n FROM pcrj_camara_servidores").fetchone()["n"]
    tot_gab = con.execute("SELECT COUNT(*) n FROM pcrj_gabinetes").fetchone()["n"]
    consultados = con.execute(
        "SELECT COUNT(DISTINCT nome_norm) n FROM pcrj_prefeitura_consulta").fetchone()["n"]
    por_conf = {r["confianca"]: r["n"] for r in con.execute(
        "SELECT confianca, COUNT(DISTINCT nome_norm) n FROM pcrj_prefeitura_consulta "
        "GROUP BY confianca")}
    vinculos = [dict(r) for r in con.execute(
        """SELECT vc.*,
                  (SELECT GROUP_CONCAT(DISTINCT s.vinculo)
                   FROM pcrj_camara_servidores s WHERE s.nome_norm=vc.nome_norm) vinculo_camara,
                  (SELECT MIN(s.ano_ingresso) FROM pcrj_camara_servidores s
                    WHERE s.nome_norm=vc.nome_norm) ingresso_camara,
                  (SELECT MIN(substr(s.data1,7,4)||'-'||substr(s.data1,4,2)||'-'||substr(s.data1,1,2))
                    FROM pcrj_camara_servidores s
                    WHERE s.nome_norm=vc.nome_norm AND s.data1 LIKE '__/__/____') ato_camara
           FROM pcrj_vinculo_cruzado vc ORDER BY vc.confianca, vc.nome_camara""")]
    # Métricas eleitorais (guardadas — a tabela pode não existir se a coleta TSE não rodou).
    cand = {"pessoas": 0, "outra_cidade": 0, "tripla": 0}
    try:
        cand["pessoas"] = con.execute(
            "SELECT COUNT(DISTINCT nome_norm) n FROM tse_candidatura").fetchone()["n"]
        cand["outra_cidade"] = con.execute(
            "SELECT COUNT(DISTINCT nome_norm) n FROM tse_candidatura WHERE outra_cidade=1").fetchone()["n"]
        cand["tripla"] = con.execute(
            """SELECT COUNT(DISTINCT tc.nome_norm) n FROM tse_candidatura tc
               WHERE EXISTS(SELECT 1 FROM pcrj_vinculo_cruzado vc
                    WHERE vc.nome_norm=tc.nome_norm AND vc.confianca='indicio_nome_unico')""").fetchone()["n"]
    except Exception as exc:
        logger.warning("métrica 'tripla' indisponível (sai zerada no relatório): %s", exc)
    return {"tot_camara": tot_camara, "tot_gab": tot_gab, "consultados": consultados,
            "por_conf": por_conf, "vinculos": vinculos, "cand": cand}


_RE_OBS = __import__("re").compile(r"admissao=(\S*)\s+exoneracao=(\S*)\s+matricula=(\S*)")


def _datas_pcrj(v: dict) -> tuple[str, str, str]:
    """(admissão, exoneração, matrícula) extraídas da observação do vínculo."""
    m = _RE_OBS.search(v.get("observacao") or "")
    if not m:
        return "", "", ""
    exo = m.group(2)
    return m.group(1), (exo if exo and exo != "None" else "—"), m.group(3)


def _ativo(v: dict) -> bool:
    """Sem data de exoneração na observação = vínculo Executivo ainda ATIVO (mais grave)."""
    return "exoneracao= " in (v.get("observacao") or "") or "exoneracao=None" in (v.get("observacao") or "")


def _tabela_vinculos(vinculos: list[dict], apenas_unico: bool, limite: int | None = None) -> str:
    conf = "indicio_nome_unico" if apenas_unico else "homonimo_ambiguo"
    sel = [v for v in vinculos if v["confianca"] == conf]
    # Ordena por FORÇA: candidato-a-acúmulo real (NÃO cessão) primeiro, depois efetivo/carreira,
    # depois vínculo ainda ativo, depois nome. Cessões descem (vínculo único, menor concern).
    sel.sort(key=lambda v: (_cessao_provavel(v), not _vinculo_efetivo(v.get("cargo_pcrj")),
                            not _ativo(v), v.get("nome_camara") or ""))
    # agrupa por PESSOA (nome_norm): variações de cargo/matrícula viram postos da mesma linha
    grupos: dict[str, list[dict]] = {}
    ordem: list[str] = []
    for v in sel:
        k = v["nome_norm"]
        if k not in grupos:
            grupos[k] = []
            ordem.append(k)
        grupos[k].append(v)
    if limite is not None:
        ordem = ordem[:limite]
    linhas = []
    for k in ordem:
        vs = grupos[k]
        v0 = vs[0]
        efet = any(_vinculo_efetivo(v.get("cargo_pcrj")) for v in vs)
        cessao = all(_cessao_provavel(v) for v in vs)
        marca = ('<span class="flag" style="background:#e8f0e8;color:#2e7d32">CESSÃO/REQUISIÇÃO</span>'
                 if cessao else ('<span class="flag">EFETIVO/CARREIRA</span>' if efet else ""))
        vistos = set()
        postos = []
        for v in vs:
            adm, exo, mat = _datas_pcrj(v)
            chave = (mat or adm, v.get("orgao_pcrj"))
            if chave in vistos:
                continue
            vistos.add(chave)
            postos.append(f"{_e(v['cargo_pcrj'])} @ {_e(v['orgao_pcrj'])}"
                          f"<br><span class='nota'>adm {_e(adm or '—')} · exo {_e(exo or '—')}"
                          f" · mat {_e(mat or '—')}</span>")
        ato = v0.get("ato_camara") or ""
        ato_fmt = f"{ato[8:10]}/{ato[5:7]}/{ato[0:4]}" if len(ato) == 10 else "—"
        desde = f"{v0.get('ingresso_camara') or '?'}<br><span class='nota'>ato {ato_fmt}</span>"
        linhas.append(
            f"<tr><td>{_e(v0['nome_camara'])} {marca}</td>"
            f"<td>{_e(v0['gabinetes'])}</td>"
            f"<td>{_e((v0.get('cargos_camara') or '')[:40])}</td>"
            f"<td>{desde}</td>"
            f"<td>{'<hr style=margin:2px>'.join(postos)}</td></tr>")
    if not linhas:
        return "<p class='nota'>Nenhum registro nesta categoria.</p>"
    return ("<table><tr><th>Nome (Câmara)</th><th>Gabinete / Vereador</th>"
            "<th>Cargo Câmara</th><th>Câmara desde</th>"
            "<th>Posto(s) na Prefeitura (admissão · exoneração · matrícula)</th></tr>"
            + "".join(linhas) + "</table>")


def _tabela_gabinetes(con) -> str:
    rows = con.execute("""
        SELECT g.gabinete_num, g.vereador,
               COUNT(DISTINCT s.nome_norm) pessoas,
               COUNT(DISTINCT vc.nome_norm) duplo
        FROM pcrj_gabinetes g
        LEFT JOIN pcrj_camara_servidores s ON s.gabinete_num=g.gabinete_num
        LEFT JOIN pcrj_vinculo_cruzado vc
             ON vc.nome_norm=s.nome_norm AND vc.confianca='indicio_nome_unico'
        GROUP BY g.gabinete_num ORDER BY duplo DESC, pessoas DESC""").fetchall()
    linhas = [f"<tr><td>Gab {r['gabinete_num']:02d}</td><td>{_e(r['vereador'])}</td>"
              f"<td style='text-align:right'>{r['pessoas']}</td>"
              f"<td style='text-align:right'>{r['duplo']}</td></tr>" for r in rows]
    return ("<table><tr><th>Gabinete</th><th>Vereador</th><th>Servidores</th>"
            "<th>Indícios de duplo vínculo</th></tr>" + "".join(linhas) + "</table>")


def _principais_achados(con) -> str:
    """Callout de abertura (padrão consultoria): nomeia os casos mais fortes de cada eixo."""
    # Eixo 1: acúmulo real (não-cessão, efetivo/carreira, concomitante).
    vinc = [dict(r) for r in con.execute(
        """SELECT vc.*, (SELECT GROUP_CONCAT(DISTINCT s.vinculo) FROM pcrj_camara_servidores s
                 WHERE s.nome_norm=vc.nome_norm) vinculo_camara
           FROM pcrj_vinculo_cruzado vc WHERE vc.confianca='indicio_nome_unico'""")]
    acumulo = [v for v in vinc if not _cessao_provavel(v) and _vinculo_efetivo(v.get("cargo_pcrj"))
               and _ativo(v)][:5]
    # Eixo 2: candidato em outra cidade com tríplice convergência, sem marca de homônimo.
    elei = con.execute("""
        SELECT tc.nome_tse, tc.cargo, tc.municipio, tc.ano,
               (SELECT COUNT(DISTINCT t2.municipio) FROM tse_candidatura t2 WHERE t2.nome_norm=tc.nome_norm) nm
        FROM tse_candidatura tc
        WHERE tc.outra_cidade=1 AND EXISTS(SELECT 1 FROM pcrj_vinculo_cruzado vc
              WHERE vc.nome_norm=tc.nome_norm AND vc.confianca='indicio_nome_unico')
        ORDER BY nm ASC, tc.ano DESC""").fetchall()
    elei = [r for r in elei if r["nm"] < 3][:5]

    li_ac = "".join(
        f"<li><b>{_e(v['nome_camara'])}</b> — {_e(v['cargo_pcrj'])} (Prefeitura, ativo) "
        f"+ {_e(v['gabinetes'] or 'quadro da Câmara')}</li>" for v in acumulo) or "<li>—</li>"
    li_el = "".join(
        f"<li><b>{_e(r['nome_tse'])}</b> — candidato a {_e(r['cargo'].lower())} em "
        f"<b>{_e(r['municipio'].title())}</b> ({r['ano']}); também Câmara + Prefeitura</li>"
        for r in elei) or "<li>—</li>"
    return (
        "<div style='border-left:4px solid #c62828;background:#fdf6f6;padding:8px 12px;margin:6px 0'>"
        "<b>A. Acúmulo concomitante mais provável</b> (efetivo/carreira no Executivo + posto na "
        f"Câmara, ativo agora — a verificar por CPF):<ul>{li_ac}</ul>"
        "<b>B. Nomeados candidatos em OUTRA cidade do RJ</b> (tríplice convergência "
        f"Câmara+Prefeitura+candidato):<ul>{li_el}</ul>"
        "<span class='nota'>Indícios para apuração — sem CPF, nome idêntico não é prova.</span></div>"
    )


def _tabela_candidaturas(con) -> str:
    """Candidaturas eleitorais dos nomeados. Ordena por FORÇA do indício:
    também-Prefeitura (tríplice convergência) → outra cidade → nome menos comum (n_munic baixo).
    Marca 'homônimo provável' quando o nome casa candidatos em ≥3 municípios distintos."""
    ver = {r["gabinete_num"]: r["vereador"] for r in con.execute(
        "SELECT gabinete_num, vereador FROM pcrj_gabinetes")}
    rows = con.execute("""
        SELECT tc.nome_tse, tc.cargo, tc.municipio, tc.ano, tc.partido, tc.outra_cidade,
               tc.nome_norm,
               (SELECT COUNT(DISTINCT t2.municipio) FROM tse_candidatura t2
                 WHERE t2.nome_norm=tc.nome_norm) n_munic,
               (SELECT COUNT(*) FROM pcrj_vinculo_cruzado vc
                 WHERE vc.nome_norm=tc.nome_norm AND vc.confianca='indicio_nome_unico') tambem_pref,
               (SELECT GROUP_CONCAT(DISTINCT s.gabinete_num) FROM pcrj_camara_servidores s
                 WHERE s.nome_norm=tc.nome_norm) gabs,
               (SELECT MIN(s.ano_ingresso) FROM pcrj_camara_servidores s
                 WHERE s.nome_norm=tc.nome_norm) ingresso
        FROM tse_candidatura tc
        ORDER BY tambem_pref DESC, tc.outra_cidade DESC, n_munic ASC, tc.nome_tse, tc.ano DESC
        """).fetchall()
    if not rows:
        return "<p class='nota'>Nenhuma candidatura casada (ou coleta TSE ainda não executada).</p>"
    linhas = []
    for r in rows:
        homon = r["n_munic"] >= 3
        gabs = [g for g in (r["gabs"].split(",") if r["gabs"] else []) if g]
        camara = "; ".join(f"Gab {g} ({ver.get(int(float(g)), '?')})" for g in gabs) or "admin/—"
        sinais = []
        if r["tambem_pref"]:
            sinais.append('<span class="flag">TAMBÉM PREFEITURA</span>')
        if r["outra_cidade"]:
            sinais.append('<span class="flag">OUTRA CIDADE</span>')
        if r["ingresso"] and r["ano"] < r["ingresso"]:
            sinais.append('<span class="flag" style="background:#fff3e0;color:#e65100">'
                          'ANTERIOR À NOMEAÇÃO</span>')
        if homon:
            sinais.append('<span class="flag" style="background:#eee;color:#777">homônimo provável</span>')
        linhas.append(
            f"<tr><td>{_e(r['nome_tse'])}</td><td>{_e(camara)} (ingresso {r['ingresso'] or '?'})</td>"
            f"<td>{_e(r['cargo'])} — {_e(r['municipio'])} ({r['ano']}, {_e(r['partido'])})</td>"
            f"<td>{' '.join(sinais)}</td></tr>")
    return ("<table><tr><th>Nome (TSE)</th><th>Vínculo Câmara (ano de ingresso)</th>"
            "<th>Candidatura (cargo — cidade, ano, partido)</th><th>Sinais</th></tr>"
            + "".join(linhas) + "</table>")


def _tabela_orgaos(con) -> str:
    rows = con.execute("""
        SELECT orgao_pcrj, COUNT(DISTINCT nome_norm) pessoas
        FROM pcrj_vinculo_cruzado WHERE confianca='indicio_nome_unico'
        GROUP BY orgao_pcrj ORDER BY pessoas DESC LIMIT 15""").fetchall()
    if not rows:
        return "<p class='nota'>Sem dados.</p>"
    linhas = [f"<tr><td>{_e(r['orgao_pcrj'])}</td>"
              f"<td style='text-align:right'>{r['pessoas']}</td></tr>" for r in rows]
    return ("<table><tr><th>Órgão da Prefeitura</th>"
            "<th>Pessoas com indício de duplo vínculo</th></tr>"
            + "".join(linhas) + "</table>")


def exportar_xlsx(destino: str, db_path=None) -> str:
    """Exporta os indícios de duplo vínculo + mapa de gabinetes para XLSX (padrão da casa)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    con = _db.conectar(db_path)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Duplo vínculo"
        cab = ["Nome (Câmara)", "Gabinete/Vereador", "Cargo Câmara", "Câmara desde (ano)",
               "Cargo Prefeitura", "Órgão Prefeitura", "Admissão PCRJ", "Exoneração PCRJ",
               "Matrícula", "Confiança", "Classificação", "Efetivo/carreira"]
        ws.append(cab)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        for v in con.execute(
                """SELECT vc.*, (SELECT GROUP_CONCAT(DISTINCT s.vinculo)
                       FROM pcrj_camara_servidores s WHERE s.nome_norm=vc.nome_norm) vinculo_camara,
                       (SELECT MIN(s.ano_ingresso) FROM pcrj_camara_servidores s
                         WHERE s.nome_norm=vc.nome_norm) ingresso_camara
                   FROM pcrj_vinculo_cruzado vc ORDER BY vc.confianca, vc.nome_camara"""):
            classe = "cessão/requisição" if _cessao_provavel(dict(v)) else "candidato a acúmulo"
            adm, exo, mat = _datas_pcrj(dict(v))
            ws.append([v["nome_camara"], v["gabinetes"], v["cargos_camara"],
                       v["ingresso_camara"], v["cargo_pcrj"],
                       v["orgao_pcrj"], adm, exo, mat, v["confianca"], classe,
                       "SIM" if _vinculo_efetivo(v["cargo_pcrj"]) else ""])
        ws3 = wb.create_sheet("Candidaturas TSE")
        ws3.append(["Nome (TSE)", "Cargo", "Município", "Ano", "Partido",
                    "Outra cidade", "Também Prefeitura", "Anterior à nomeação",
                    "Homônimo provável"])
        for c in ws3[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        try:
            for t in con.execute("""
                SELECT tc.*,
                   (SELECT COUNT(DISTINCT t2.municipio) FROM tse_candidatura t2
                     WHERE t2.nome_norm=tc.nome_norm) n_munic,
                   (SELECT COUNT(*) FROM pcrj_vinculo_cruzado vc
                     WHERE vc.nome_norm=tc.nome_norm AND vc.confianca='indicio_nome_unico') pref,
                   (SELECT MIN(s.ano_ingresso) FROM pcrj_camara_servidores s
                     WHERE s.nome_norm=tc.nome_norm) ingresso
                FROM tse_candidatura tc ORDER BY pref DESC, tc.outra_cidade DESC, tc.nome_tse"""):
                ws3.append([t["nome_tse"], t["cargo"], t["municipio"], t["ano"], t["partido"],
                            "SIM" if t["outra_cidade"] else "", "SIM" if t["pref"] else "",
                            "SIM" if (t["ingresso"] and t["ano"] < t["ingresso"]) else "",
                            "SIM" if t["n_munic"] >= 3 else ""])
        except Exception as exc:
            logger.warning("aba de candidaturas do xlsx falhou (sai incompleta): %s", exc)
        ws2 = wb.create_sheet("Gabinetes")
        ws2.append(["Gabinete", "Vereador"])
        for g in con.execute("SELECT gabinete_num, vereador FROM pcrj_gabinetes ORDER BY gabinete_num"):
            ws2.append([g["gabinete_num"], g["vereador"]])
        wb.save(destino)
    finally:
        con.close()
    return destino


def montar_ctx(db_path=None) -> dict:
    con = _db.conectar(db_path)
    try:
        d = _coletar(con)
        unico = d["por_conf"].get("indicio_nome_unico", 0)
        ambiguo = d["por_conf"].get("homonimo_ambiguo", 0)
        nao = d["por_conf"].get("nao_encontrado", 0)
        indisp = d["por_conf"].get("indisponivel", 0)
        unicos_v = [v for v in d["vinculos"] if v["confianca"] == "indicio_nome_unico"]
        # unidade = PESSOA (nome_norm), nunca linha: variações de cargo/matrícula não inflam contagem
        pessoas: dict[str, list[dict]] = {}
        for v in unicos_v:
            pessoas.setdefault(v["nome_norm"], []).append(v)
        # pessoa é "cessão" só se TODOS os postos são cessão; senão é candidata a acúmulo real
        cessoes = sum(1 for vs in pessoas.values() if all(_cessao_provavel(v) for v in vs))
        candidatos_p = [vs for vs in pessoas.values() if not all(_cessao_provavel(v) for v in vs)]
        efetivos = sum(1 for vs in candidatos_p if any(_vinculo_efetivo(v.get("cargo_pcrj")) for v in vs))
        # Concomitância: algum vínculo no Executivo ainda ATIVO (sem exoneração na consulta) →
        # duplo vínculo CONCORRENTE agora (quadro da Câmara é o roster atual). Sinal mais forte.
        concomitantes = sum(1 for vs in candidatos_p if any(_ativo(v) for v in vs))
        n_cand = len(candidatos_p)
        # Índice de convergência (indicador interno, NÃO punição): concentra nos candidatos reais.
        score = min(100, n_cand * 1 + concomitantes * 2 + efetivos * 3)
        faixa = "ALTO" if concomitantes >= 20 or efetivos >= 15 else (
                "MÉDIO" if n_cand >= 10 else "BAIXO")

        janela_cam = con.execute(
            "SELECT MAX(substr(coletado_em,1,10)) FROM pcrj_camara_servidores").fetchone()[0] or "?"
        janela_pref = con.execute(
            "SELECT MAX(substr(consultado_em,1,10)) FROM pcrj_prefeitura_consulta").fetchone()[0] or "?"
        nota_temporal = (
            f"<p class='nota'><b>Janela temporal:</b> Câmara = relação ATUAL de servidores "
            f"(dados abertos, coletada em {janela_cam}; quem já saiu não consta). Prefeitura = consulta ao "
            f"contracheque em {janela_pref}; <b>ATIVO</b> = sem exoneração <i>nessa consulta</i>. "
            f"Contagens por PESSOA (nome único), não por linha de cargo/matrícula. "
            f"Datas comparadas cronologicamente. Nome idêntico sem CPF = indício, não prova.</p>")

        sumario = (
            nota_temporal +
            f"<table>"
            f"<tr><td>Servidores da Câmara analisados (nomes distintos)</td>"
            f"<td style='text-align:right'><b>{d['tot_camara']:,}</b></td></tr>"
            f"<tr><td>Gabinetes parlamentares mapeados</td>"
            f"<td style='text-align:right'>{d['tot_gab']}</td></tr>"
            f"<tr><td>Nomes consultados na Prefeitura</td>"
            f"<td style='text-align:right'>{d['consultados']:,}</td></tr>"
            f"<tr><td>Indícios de duplo vínculo (nome único)</td>"
            f"<td style='text-align:right'>{unico}</td></tr>"
            f"<tr><td>&nbsp;&nbsp;(−) prováveis cessões/requisições (à disposição da CMRJ — vínculo ÚNICO)</td>"
            f"<td style='text-align:right'>{cessoes}</td></tr>"
            f"<tr><td><b>&nbsp;&nbsp;(=) candidatos a acúmulo REAL (dois postos distintos)</b></td>"
            f"<td style='text-align:right'><b>{n_cand}</b></td></tr>"
            f"<tr><td>&nbsp;&nbsp;&nbsp;&nbsp;dos quais efetivo/carreira no Executivo</td>"
            f"<td style='text-align:right'>{efetivos}</td></tr>"
            f"<tr><td>&nbsp;&nbsp;&nbsp;&nbsp;<b>com vínculo no Executivo ATIVO (concomitância provável)</b></td>"
            f"<td style='text-align:right'><b>{concomitantes}</b></td></tr>"
            f"<tr><td>Homônimos ambíguos (não isoláveis sem CPF)</td>"
            f"<td style='text-align:right'>{ambiguo}</td></tr>"
            f"<tr><td>Sem correspondência na Prefeitura</td>"
            f"<td style='text-align:right'>{nao:,}</td></tr>"
            f"<tr><td colspan='2' style='background:#f0f4fa'><b>Cruzamento eleitoral (TSE)</b></td></tr>"
            f"<tr><td>Nomeados que foram <b>candidatos</b> em alguma eleição</td>"
            f"<td style='text-align:right'><b>{d['cand']['pessoas']}</b></td></tr>"
            f"<tr><td>&nbsp;&nbsp;<b>candidatos em OUTRA cidade</b> (flag reforçada)</td>"
            f"<td style='text-align:right'><b>{d['cand']['outra_cidade']}</b></td></tr>"
            f"<tr><td>&nbsp;&nbsp;tríplice convergência (Câmara + Prefeitura + candidatura)</td>"
            f"<td style='text-align:right'>{d['cand']['tripla']}</td></tr>"
            f"<tr><td>Indisponível (erro de consulta)</td>"
            f"<td style='text-align:right'>{indisp}</td></tr>"
            f"</table>".replace(",", ".")
        )

        metodologia = (
            "<p>Cruzamento <b>direcional</b>: a relação completa de servidores da Câmara "
            "(dados abertos, por ano de ingresso) foi confrontada, nome a nome, com a "
            "consulta de remuneração da Prefeitura (contrachequeapi.rio.gov.br) em "
            "competências-amostra (gestão atual + 06/2024 + 06/2021).</p>"
            "<p><b>Limitações honestas:</b> (1) nenhuma das bases públicas expõe CPF → o "
            "casamento é por <b>nome normalizado idêntico</b>, que é INDÍCIO e não prova "
            "(homônimo possível); (2) a busca da Prefeitura é por substring do nome; "
            "(3) só as competências-amostra foram varridas — quem passou pelo Executivo "
            "fora delas pode não aparecer; (4) 'nome único' = exatamente 1 matrícula na "
            "Prefeitura com o nome (mais forte); 'homônimo ambíguo' = 2+ matrículas "
            "(não isolável). A confirmação exige cruzamento por CPF na origem (RH).</p>"
            "<p><b>Cessão ≠ acúmulo (nuance decisiva):</b> quando o servidor da Prefeitura está "
            "'à disposição da CMRJ' ou a Câmara o lista como 'Requisitado', as duas folhas "
            "retratam o MESMO posto (cessão/requisição, vínculo único) — NÃO é acúmulo. Esses "
            "casos são marcados e descontados; os <b>candidatos a acúmulo real</b> são os que "
            "ocupam dois postos distintos (ex.: comissionado na Câmara + efetivo operacional no "
            "Executivo).</p>"
            "<p><b>Base legal do alerta:</b> acumulação remunerada de cargos públicos é "
            "vedada salvo exceções constitucionais (CF art. 37, XVI/XVII). Guarda Municipal, "
            "professor e demais cargos efetivos/carreira no Executivo sobrepostos a posto "
            "comissionado no Legislativo são o indício mais forte — a apurar, jamais acusação.</p>"
            "<p><b>Cruzamento eleitoral (TSE):</b> os nomes foram confrontados com as candidaturas "
            "do TSE (dados abertos, arquivo do estado do RJ) de 2016/2020/2024 (municipais) e "
            "2018/2022 (gerais). CPF mascarado → match por nome. Como nome comum gera homônimo, "
            "marcamos 'homônimo provável' quando um nome casa candidatos em ≥3 municípios distintos. "
            "Sinal mais forte = <b>tríplice convergência</b> (nome na Câmara + Prefeitura + candidato); "
            "<b>flag reforçada</b> = candidatura em município ≠ Rio de Janeiro. Limitação: candidatura "
            "em OUTRO estado não é captada nesta v1 (apenas o arquivo do RJ).</p>"
        )

        prov = [
            {"dado": "Servidores Câmara RJ", "estado": "REAL",
             "fonte": "transparencia.camara.rj.gov.br (dados abertos)",
             "data": datetime.now(timezone.utc).strftime("%d/%m/%Y")},
            {"dado": "Gabinete→Vereador", "estado": "REAL",
             "fonte": "CMRJ — tabela de núcleos dos gabinetes (.xls)",
             "data": datetime.now(timezone.utc).strftime("%d/%m/%Y")},
            {"dado": "Remuneração Prefeitura RJ", "estado": "REAL",
             "fonte": "contrachequeapi.rio.gov.br (consulta pública)",
             "data": datetime.now(timezone.utc).strftime("%d/%m/%Y")},
        ]

        secoes = [
            {"titulo": "Principais achados", "html": _principais_achados(con)},
            {"titulo": "1. Sumário executivo", "html": sumario},
            {"titulo": "2. Indícios de duplo vínculo Câmara × Prefeitura (nome único)",
             "html": _tabela_vinculos(d["vinculos"], apenas_unico=True)},
            {"titulo": "3. Homônimos ambíguos (a verificar por CPF — não isoláveis)",
             "html": ("<p class='nota'>Amostra dos primeiros 80 nomes no PDF; a íntegra "
                      "está na planilha (xlsx) que acompanha este relatório.</p>"
                      + _tabela_vinculos(d["vinculos"], apenas_unico=False, limite=80))},
            {"titulo": "4. Concentração por órgão da Prefeitura",
             "html": _tabela_orgaos(con)},
            {"titulo": "5. Nomeados que foram candidatos em eleições (TSE) — "
                       "flag reforçada: candidatura em OUTRA cidade",
             "html": _tabela_candidaturas(con)},
            {"titulo": "6. Mapa por gabinete (servidores e indícios)",
             "html": _tabela_gabinetes(con)},
            {"titulo": "7. Metodologia e limitações", "html": metodologia},
        ]
        return {
            "classificacao": "CONFIDENCIAL — CONTROLE EXTERNO",
            "titulo": "Cruzamento Câmara Municipal × Prefeitura do Rio de Janeiro",
            "subtitulo": "Indícios de duplo vínculo por nome e mapa de gabinetes — Módulo PCRJ v1",
            "metodologia": "Cruzamento nominal direcional + níveis de confiança",
            "score": score, "faixa": faixa,
            "top_flags": [f"{n_cand} candidatos acúmulo", f"{concomitantes} concomitantes",
                          f"{cessoes} cessões (vínculo único)"],
            "secoes": secoes, "proveniencia": prov,
            "ressalva": "Indícios para apuração por CPF na origem (RH). Presunção de "
                        "legitimidade dos atos; INDISPONÍVEL não é ausência de vínculo.",
            "_dados": {"unico": unico, "efetivos": efetivos, "ambiguo": ambiguo},
            "cand": d["cand"],
        }
    finally:
        con.close()


async def gerar(db_path=None) -> dict:
    """Gera md + pdf do relatório. Retorna caminhos."""
    from compliance_agent.reporting.render_html import render_html, html_to_pdf
    from pathlib import Path
    ctx = montar_ctx(db_path)
    html = render_html(ctx)
    base = Path(__file__).resolve().parents[2] / "reports"
    base.mkdir(exist_ok=True)
    data = datetime.now().date()
    pdf = str(base / f"pcrj_camara_cruzamento_{data}.pdf")
    htmlp = str(base / f"pcrj_camara_cruzamento_{data}.html")
    xlsx = str(base / f"pcrj_camara_cruzamento_{data}.xlsx")
    Path(htmlp).write_text(html, encoding="utf-8")
    exportar_xlsx(xlsx, db_path)
    await html_to_pdf(html, pdf)
    return {"pdf": pdf, "html": htmlp, "xlsx": xlsx, "score": ctx["score"], "faixa": ctx["faixa"]}


if __name__ == "__main__":
    import asyncio
    print(asyncio.run(gerar()))
