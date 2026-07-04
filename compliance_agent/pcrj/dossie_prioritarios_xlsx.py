# -*- coding: utf-8 -*-
"""XLSX dos alvos prioritários (para distribuir à equipe da CPI).

Uma planilha, uma linha por alvo 'forte', colunas separadas por indício + coluna de
"próximo passo" e status de apuração. Reusa a tabela `pcrj_fantasma_servidor`.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from compliance_agent.pcrj import db as _db

_CAB = ["#", "Nome", "Gabinete(s)", "Cargo Câmara", "Score", "Faixa",
        "Acúmulo Câmara∩Prefeitura", "Benefício (BF/BPC)", "Domicílio distante",
        "Candidato outra cidade", "Homônimo?", "Indícios (texto)",
        "Próximo passo (CPI)", "Status apuração"]


def _flags(sinais: str) -> dict:
    s = sinais or ""
    return {
        "acumulo": "SIM" if "acúmulo" in s.lower() else "",
        "beneficio": "🔴 " + ("BPC" if "BPC" in s else "") + (" Bolsa Família" if "Bolsa Família" in s else "")
        if "recebe" in s else "",
        "distante": "SIM" if "domicílio eleitoral" in s.lower() else "",
        "candidato": "SIM" if "candidato em outra" in s.lower() else "",
    }


def gerar(db_path=None) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    con = _db.conectar(db_path)
    rows = con.execute(
        "SELECT nome, gabinetes, cargos_camara, sinais, score, faixa, homonimo "
        "FROM pcrj_fantasma_servidor WHERE faixa='forte' ORDER BY score DESC, nome").fetchall()
    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Alvos prioritários"
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True)
    ws.append(_CAB)
    for c in ws[1]:
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    passo = ("Requisição CPI: CPF+endereço (Receita/RH) · ponto/frequência · eSocial/CAGED · "
             "consulta manual TJRJ/TRT1 por nome")
    for i, r in enumerate(rows, 1):
        f = _flags(r["sinais"])
        ws.append([i, r["nome"], r["gabinetes"], r["cargos_camara"], r["score"], r["faixa"],
                   f["acumulo"], f["beneficio"], f["distante"], f["candidato"],
                   "provável" if r["homonimo"] else "", r["sinais"], passo, ""])

    larguras = [4, 30, 22, 26, 6, 10, 14, 16, 12, 12, 10, 60, 42, 16]
    for col, w in enumerate(larguras, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(38 + col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(rows) + 1}"

    base = Path(__file__).resolve().parents[2] / "reports"
    base.mkdir(exist_ok=True)
    out = str(base / f"pcrj_alvos_prioritarios_{datetime.now().date()}.xlsx")
    wb.save(out)
    return {"xlsx": out, "total": len(rows)}


if __name__ == "__main__":
    import json
    print(json.dumps(gerar(), ensure_ascii=False))
