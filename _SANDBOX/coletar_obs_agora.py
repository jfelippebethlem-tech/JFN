#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MGS CLEAN — Coleta de Ordens Bancárias SIAFE (2023–2026)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

RODA NA MÁQUINA COM ACESSO AO SIAFE (Windows local ou VM Oracle).
Este servidor cloud NÃO tem acesso direto — WAF bloqueia IP não-governamental.

Uso simples (Windows CMD ou PowerShell):
    python _SANDBOX/coletar_obs_agora.py

Credenciais em ~/.hermes/.env ou JFN/.env:
    SIAFE_USER=<CPF>
    SIAFE_PASS=<senha>

Saída automática:
    data/sei_cache/mgsclean_obs_AAAA.json   — por ano
    data/sei_cache/mgsclean_obs_todas.json  — consolidado
    data/sei_cache/mgsclean_obs_resumo.md   — relatório Markdown

Depois: git add + git commit + git push (automático ao final).
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

# Habilita flush imediato de stdout (necessário para logs em tempo real no GitHub Actions)
sys.stdout.reconfigure(line_buffering=True)

# ── Carregar .env ──────────────────────────────────────────────────────────────
for _env_path in [
    Path.home() / ".hermes" / ".env",
    Path(__file__).parents[1] / ".env",
    Path("C:/JFN/jfn/.env"),
    Path("C:/Users") / os.environ.get("USERNAME","user") / "JFN" / ".env",
]:
    if _env_path.exists():
        for _ln in _env_path.read_text(encoding="utf-8", errors="replace").splitlines():
            _ln = _ln.strip()
            if _ln and not _ln.startswith("#") and "=" in _ln:
                k, v = _ln.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

# ── Constantes ─────────────────────────────────────────────────────────────────
CNPJ        = "19088605000104"
CNPJ_FMT    = "19.088.605/0001-04"
NOME_EMP    = "MGS CLEAN SOLUCOES E SERVICOS LTDA"
LOGIN_URL   = "https://siafe2.fazenda.rj.gov.br/Siafe/faces/login.jsp"
REPO_ROOT   = Path(__file__).parents[1]
CACHE_DIR   = REPO_ROOT / "data" / "sei_cache"
SS_DIR      = REPO_ROOT / "screenshots" / "obs_coleta"
TIMEOUT     = 40_000

# Exercício → valor do SELECT de login (confirmado ao vivo 2026-06-05)
EXERCICIOS = {2027: "0", 2026: "1", 2025: "2", 2024: "3", 2023: "4"}

MESES_PT = {
    1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
    7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro",
}

# Nomes de UG conhecidos — Secretarias e órgãos do Estado do RJ
# Fonte: SIAFE-Rio 2 / Portal da Transparência RJ
_UG_NOMES: dict[str, str] = {
    # Poder Executivo — Secretarias
    "270001": "Casa Civil",
    "270002": "SEFAZ — Sec. Fazenda",
    "270003": "FUNESBOM",
    "270004": "SESDEC — Sec. Defesa Civil",
    "270005": "SEAP — Sec. Administração Penitenciária",
    "270006": "TCE-RJ",
    "270007": "SESEG — Sec. Segurança",
    "270008": "SSP — Subsec. Segurança",
    "270009": "PGE — Proc. Geral Estado",
    "270010": "SEAAPI — Sec. Agricultura",
    "270011": "SEEDUC — Sec. Educação",
    "270012": "SECCR — Sec. Esporte",
    "270013": "SES — Sec. Saúde",
    "270014": "SEST — Sec. Trabalho",
    "270015": "SECEC — Sec. Cultura",
    "270016": "FUNESBOM (FEE)",
    "270017": "DETRAN-RJ",
    "270018": "DER-RJ",
    "270019": "SEHAB — Sec. Habitação",
    "270020": "RIOPREVIDÊNCIA",
    "270021": "SEOBRAS — Sec. Obras",
    "270022": "SEOP — Sec. Obras Públicas",
    "270023": "SEMADS — Sec. Meio Ambiente",
    "270024": "INEA — Inst. Amb. Est.",
    "270025": "SETUR — Sec. Turismo",
    "270026": "SEIO — Sec. Interior",
    "270027": "SECI — Sec. Ciência",
    "270028": "SERLA — Fund. Rio Águas",
    "270029": "FES — Fundo Estadual Saúde",
    "270030": "PRODERJ — TI Estado",
    "270031": "SEJUDH — Sec. Direitos Humanos",
    "270032": "SEDS — Sec. Desenvolvimento Social",
    "270033": "SEFAZ — DG",
    "270034": "SEDAM — Sec. Desenv. Agro.",
    "270035": "SEPLAG — Planejamento",
    "270036": "SDE — Sec. Desenv. Econômico",
    "270037": "SECTMA — Ciência e Tecnologia",
    "270038": "SETRANS — Sec. Transportes",
    "270039": "SEINFRA — Infraestrutura",
    "270040": "SEEA — Sec. Energia",
    "270041": "SESMICT — Inovação",
    "270042": "ITERJ — Instituto Terras",
    "270043": "EMATER-Rio",
    "270044": "SECOM — Comunicação",
    "270045": "SEJURIS — Jurídico",
    "270046": "SECTUR",
    "270047": "SEAS — Assist. Social",
    "270048": "SESM — Serviço Militar",
    "270049": "SEEQ",
    "270050": "SEPEN — Penitenciária",
    "270051": "PMERJ — Polícia Militar",
    "270052": "PCERJ — Polícia Civil",
    "270053": "CBMERJ — Bombeiros",
    "270054": "IGP-RJ",
    "270060": "Casa Civil (DG)",
    "270070": "ALERJ — Assembléia Leg.",
    "270075": "TCE-RJ (Tribunal)",
    "270080": "FAPERJ",
    "270081": "UERJ",
    "270082": "UENF",
    "270083": "UEZO",
    "270084": "FAETEC",
    "270085": "CEFET-RJ",
    "270086": "FENORTE-RJ",
    "270090": "CEP-RJ",
    "270091": "CREA-RJ",
    "270092": "CODIN",
    "270093": "INVESTE-RIO",
    "270094": "CEDAE",
    "270095": "LIGHT (Estado)",
    "270096": "FLUMITRENS",
    "270097": "METRO-RIO (Estado)",
    "270100": "SuperVia (Estado)",
    "300100": "SEFAZ-RJ — Receita",
    "300200": "Tesouro Estadual",
    "300300": "FESP — Fund. Escola Serv.",
    "320001": "Fundo Especial Saúde",
    "320002": "Fundo Manutenção Educ.",
    "510001": "DER — Fundo Rodoviário",
    "510002": "RIOPREVIDÊNCIA (Fundo)",
    "510003": "FES (Fundo Saúde)",
    "510004": "FECP",
}

def _ug_nome(c: str) -> str:
    return _UG_NOMES.get(str(c).strip(), str(c))


async def _descobrir_ugs_siafe(page) -> list[str]:
    """
    Descobre UGs disponíveis na tela de OBs do SIAFE (campo de filtro UG emitente).
    Retorna lista de códigos de UG. Fallback: lista estática _UG_NOMES.
    """
    print("  → Descobrindo UGs disponíveis no SIAFE …", flush=True)
    ugs_found = await page.evaluate("""() => {
        const ugs = new Set();
        // 1. Procura em selects com "ug", "emitente", "orgao" no ID ou label
        for (const sel of document.querySelectorAll('select')) {
            const id  = (sel.id   || '').toLowerCase();
            const lbl = (document.querySelector('label[for="' + sel.id + '"]') || {}).textContent || '';
            const ltl = lbl.toLowerCase();
            if (id.includes('ug') || id.includes('emitente') || id.includes('orgao') ||
                ltl.includes('ug') || ltl.includes('emitente') || ltl.includes('unidade gestora')) {
                for (const opt of sel.options) {
                    const v = opt.value.trim();
                    if (v && /^\\d{5,6}$/.test(v)) ugs.add(v);
                }
            }
        }
        // 2. Procura em qualquer option com valor numérico de 6 dígitos (padrão UG SIAFE)
        if (ugs.size < 3) {
            for (const opt of document.querySelectorAll('select option')) {
                const v = opt.value.trim();
                if (/^\\d{6}$/.test(v)) ugs.add(v);
            }
        }
        return [...ugs].slice(0, 300);
    }""")

    if len(ugs_found) >= 3:
        print(f"  ✔ {len(ugs_found)} UGs descobertas no SIAFE (interface)")
        # Mescla com lista estática para completar nomes
        for ug in ugs_found:
            _UG_NOMES.setdefault(ug, ug)
        return ugs_found

    # Fallback: lista estática
    ugs_static = list(_UG_NOMES.keys())
    print(f"  → UGs estáticas (fallback): {len(ugs_static)}")
    return ugs_static


async def _aplicar_filtro_ug(pg, ug_code: str) -> bool:
    """Aplica filtro de UG emitente na tela de OBs. Retorna True se conseguiu."""
    print(f"  → Filtrando por UG {ug_code} ({_ug_nome(ug_code)}) …", flush=True)

    # Abre acordeão de filtro
    for acc_id in [
        "pt1:tblOrdemBancaria:pnlAccordionDec_afrCl0",
        "pt1:tblOrdemBancaria:sdtFilter::disAcr",
    ]:
        try:
            loc = pg.locator(f'[id*="{acc_id.split(":")[-1]}"]').first
            if await loc.count() > 0:
                await loc.click(timeout=4000, force=True)
                await asyncio.sleep(2)
                break
        except Exception:
            pass

    filled = await pg.evaluate("""(ug) => {
        const candidatos = [...document.querySelectorAll(
            'input[type="text"], input:not([type])'
        )].filter(el => {
            const id  = (el.id  || '').toLowerCase();
            const ph  = (el.placeholder || '').toLowerCase();
            const lbl = document.querySelector('label[for="' + el.id + '"]');
            const lt  = lbl ? lbl.textContent.toLowerCase() : '';
            return (id.includes('emitente') || id.includes('ug') ||
                    ph.includes('ug') || lt.includes('ug emitente') ||
                    lt.includes('unidade gestora'))
                   && el.getBoundingClientRect().width > 0;
        });
        if (!candidatos.length) return null;
        const el = candidatos[0];
        el.focus(); el.value = ug;
        ['input','change','blur','keyup'].forEach(e =>
            el.dispatchEvent(new Event(e, {bubbles: true})));
        return el.id || 'ok';
    }""", ug_code)

    if not filled:
        return False

    await asyncio.sleep(1.5)
    # Clica Pesquisar
    await pg.evaluate("""() => {
        for (const el of document.querySelectorAll('button, a, input[type="button"]')) {
            const t = (el.textContent || el.value || '').trim().toLowerCase();
            if ((t.includes('pesquis') || t.includes('filtrar') || t.includes('consult')) &&
                el.getBoundingClientRect().width > 0) { el.click(); return t; }
        }
    }""")
    await _settle(pg, 6000)
    await _dismiss_popups(pg)
    return True

def _brl(v: float) -> str:
    s = f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    return f"R$ {s}"

def _load_empresas() -> list[dict]:
    """Carrega lista de empresas alvo do arquivo de configuração."""
    env_cnpjs = os.environ.get("SIAFE_CNPJS", "").strip()
    if env_cnpjs:
        empresas = []
        for raw in env_cnpjs.split(","):
            cnpj = re.sub(r"[^\d]", "", raw.strip())
            if len(cnpj) == 14:
                fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                empresas.append({"cnpj": cnpj, "cnpj_fmt": fmt, "nome": cnpj, "categoria": f"obs_{cnpj}"})
        if empresas:
            return empresas

    cfg = REPO_ROOT / "data" / "empresas_target.json"
    if cfg.exists():
        try:
            dados = json.loads(cfg.read_text(encoding="utf-8"))
            return dados if isinstance(dados, list) else dados.get("empresas", [])
        except Exception:
            pass

    return [{"cnpj": CNPJ, "cnpj_fmt": CNPJ_FMT, "nome": NOME_EMP, "categoria": "mgs_clean_real"}]

def _parse_money(s) -> float:
    try:
        return float(re.sub(r"[^\d,]","",str(s)).replace(",",".") or "0")
    except Exception:
        return 0.0

def _parse_date(s: str) -> Optional[datetime]:
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except Exception:
            pass
    return None


# ── Helpers Telegram ──────────────────────────────────────────────────────────

def _telegram(msg: str):
    """Envia notificação Telegram (silencioso se falhar)."""
    try:
        import urllib.request, urllib.parse
        token   = os.environ.get("TELEGRAM_BOT_TOKEN","")
        chat_id = os.environ.get("TELEGRAM_CHAT_ID","")
        if not token or not chat_id:
            return
        url  = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": msg[:4000]}).encode()
        urllib.request.urlopen(url, data=data, timeout=10)
    except Exception:
        pass


async def _poll_github_for_mfa(repo: str, masked_email: str, ano: int = 0) -> str:
    """
    Polling do arquivo data/sei_cache/.mfa_input no GitHub (branch feature).
    O arquivo é escrito externamente (por IA/humano via push) com o código MFA.
    Script permanece na MESMA sessão SIAFE — não faz novo login.

    Para suportar múltiplos exercícios (cada um com seu próprio MFA):
    - Lê o SHA atual do arquivo ANTES de começar o wait
    - Só aceita código com SHA DIFERENTE do SHA inicial (novo push)
    - Assim nunca reutiliza código de exercício anterior
    """
    import urllib.request, base64, time as _time

    branch   = os.environ.get("GITHUB_HEAD_REF") or "claude/rj-finance-agent-BYlhJ"
    mfa_path = "data/sei_cache/.mfa_input"
    api_url  = f"https://api.github.com/repos/{repo}/contents/{mfa_path}"
    loop     = asyncio.get_event_loop()

    def _fetch_file(ts: int):
        req = urllib.request.Request(
            f"{api_url}?ref={branch}&t={ts}",
            headers={"Accept": "application/vnd.github+json", "User-Agent": "siafe-collector/2.0"},
        )
        return json.loads(urllib.request.urlopen(req, timeout=8).read())

    # Lê SHA atual (para ignorar código já existente de exercício anterior)
    initial_sha = ""
    try:
        current = await loop.run_in_executor(None, lambda: _fetch_file(0))
        initial_sha = current.get("sha", "")
        existing   = base64.b64decode(current.get("content","")).decode("utf-8").strip()
        if existing:
            print(f"  → MFA polling: arquivo tem conteúdo antigo (sha={initial_sha[:7]}) — aguardando novo push")
    except Exception:
        pass

    ano_str = f" exercício {ano}" if ano else ""
    print(f"\n  ╔══════════════════════════════════════════════════════════╗")
    print(f"  ║  AGUARDANDO CÓDIGO MFA{ano_str:<35}║")
    print(f"  ║  SIAFE enviou código para: {masked_email[:28]:<28}║")
    print(f"  ║  Empurre o código para o GitHub:                        ║")
    print(f"  ║  Arquivo: {mfa_path:<48}║")
    print(f"  ║  Branch : {branch:<48}║")
    print(f"  ║  Conteúdo: apenas o código alfanumérico (ex: aB3xYz)   ║")
    print(f"  ╚══════════════════════════════════════════════════════════╝\n")

    # Notifica via Telegram se configurado
    _telegram(f"🔐 SIAFE MFA{ano_str}\nCódigo enviado para: {masked_email}\nEmpurre para GitHub:\n{mfa_path} (branch {branch})")

    deadline = _time.time() + 300
    last_sha = initial_sha  # Só aceita SHAs novos (diferentes do inicial)

    while _time.time() < deadline:
        await asyncio.sleep(10)
        try:
            data = await loop.run_in_executor(None, lambda: _fetch_file(int(_time.time())))
            sha  = data.get("sha", "")
            raw  = base64.b64decode(data.get("content", "")).decode("utf-8").strip()
            if raw and sha != last_sha and re.match(r"^[A-Za-z0-9]{4,12}$", raw):
                last_sha = sha
                print(f"  → Código MFA recebido via GitHub polling: ******")
                _telegram(f"✅ Código MFA{ano_str} recebido — autenticando no SIAFE")
                return raw
        except Exception as _e:
            if "404" not in str(_e):
                pass
        remaining = int(deadline - _time.time())
        if remaining > 0 and remaining % 60 == 0:
            print(f"  → Aguardando código MFA… {remaining}s restantes")
            _telegram(f"⏳ SIAFE MFA{ano_str}: {remaining}s restantes")

    print("  → Timeout GitHub polling (5 min)")
    _telegram(f"❌ SIAFE MFA{ano_str}: timeout — código não recebido em 5 min")
    return ""


async def _aguardar_codigo_mfa(masked_email: str, ano: int = 0) -> str:
    """
    SIAFE tem MFA — verifica env SIAFE_MFA_CODE primeiro, depois GitHub polling, depois Telegram.
    Retorna o código ou "" se não disponível.
    """
    # Código fornecido diretamente via env var (run com MFA code pré-configurado)
    direct = os.environ.get("SIAFE_MFA_CODE", "").strip()
    if direct:
        print(f"  → Código MFA via env SIAFE_MFA_CODE: ******")
        return direct

    import time
    import urllib.request

    # GitHub file polling (CI/cloud — mantém sessão SIAFE ativa, não faz novo login)
    gh_repo = os.environ.get("GITHUB_REPOSITORY", "")
    if gh_repo and (os.environ.get("CI") or os.environ.get("GITHUB_ACTIONS")):
        code = await _poll_github_for_mfa(gh_repo, masked_email, ano=ano)
        if code:
            return code
        # Polling timeout — fall through to Telegram if configured

    token   = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    if not token or not chat_id:
        print("  → MFA: sem GitHub repo, sem Telegram — configure SIAFE_MFA_CODE ou TELEGRAM_BOT_TOKEN")
        return ""

    _telegram(
        f"🔐 SIAFE MFA — coletar OBs MGS CLEAN\n"
        f"Código enviado para: {masked_email}\n"
        f"Responda esta mensagem com o código de 6 dígitos."
    )
    print(f"  → MFA: aguardando código via Telegram (máx 5 min) …")

    # Obtém último update_id para não pegar mensagens antigas
    last_id = 0
    loop = asyncio.get_event_loop()
    try:
        r = await loop.run_in_executor(None, lambda: urllib.request.urlopen(
            f"https://api.telegram.org/bot{token}/getUpdates?limit=1&offset=-1", timeout=10
        ))
        data = json.loads(r.read())
        if data.get("result"):
            last_id = data["result"][-1]["update_id"]
    except Exception as e:
        print(f"    [w] getUpdates inicial: {e}")

    deadline = time.time() + 300
    while time.time() < deadline:
        await asyncio.sleep(5)
        try:
            r = await loop.run_in_executor(None, lambda: urllib.request.urlopen(
                f"https://api.telegram.org/bot{token}/getUpdates?offset={last_id+1}&timeout=5",
                timeout=8,
            ))
            data = json.loads(r.read())
            for upd in data.get("result", []):
                last_id = upd["update_id"]
                txt = ((upd.get("message") or {}).get("text") or "").strip()
                if re.match(r"^[A-Za-z0-9]{4,12}$", txt):
                    print(f"  → Código MFA recebido: ******")
                    return txt
        except Exception:
            pass

    print("  → Timeout MFA (5 min) — login cancelado")
    return ""


# ── Helpers Browser ───────────────────────────────────────────────────────────

async def _settle(pg, ms: int = 4000):
    try:
        await pg.wait_for_load_state("networkidle", timeout=ms)
    except Exception:
        pass
    await asyncio.sleep(max(1.5, ms / 2500))


async def _dismiss_popups(pg, tries: int = 6, allow_confirm: bool = False):
    """Fecha modais ADF que travam a navegação.

    allow_confirm=True: clica em confirmar/continuar/sim em qualquer popup (PRÉ-login).
    allow_confirm=False: apenas 'ok/sim/fechar', evita re-click no btnConfirmar de login.
    """
    kws = ['ok','sim','fechar','confirmar','continuar','continue','prosseguir','yes'] \
          if allow_confirm else ['ok','sim','fechar']
    # Sempre exclui o botão principal de submit do formulário de login para não entrar em loop.
    skip_ids = ['btnconfirmar']

    for _ in range(tries):
        result = await pg.evaluate(f"""() => {{
            const kws     = {json.dumps(kws)};
            const skipIds = {json.dumps(skip_ids)};

            // Verifica visibilidade via rect OU offset (cobre dialogs ADF animados)
            const isVisible = el => {{
                const r = el.getBoundingClientRect();
                return (r.width > 0 && r.height > 0) || (el.offsetWidth > 0 && el.offsetHeight > 0);
            }};

            // 1. Tenta em overlays/dialogs ADF primeiro (selectors específicos)
            for (const root of document.querySelectorAll(
                '[id*="popup"], [id*="dlg"], [id*="modal"], [id*="dialog"], ' +
                '[role="dialog"], [role="alertdialog"], .xc9, .x1n, .AFPanelWindow'
            )) {{
                for (const el of root.querySelectorAll(
                    'button, a, input[type="button"], input[type="submit"]'
                )) {{
                    const t = (el.textContent || el.value || '').trim().toLowerCase();
                    const id = (el.id || '').toLowerCase();
                    if (isVisible(el) && kws.some(k => t === k || t.startsWith(k + ' '))
                        && !skipIds.some(s => id.includes(s))) {{
                        el.click();
                        return 'overlay:' + t + '[' + el.id + ']';
                    }}
                }}
            }}

            // 2. Fallback: qualquer botão visível — exclui apenas btnConfirmar quando pedido
            for (const el of document.querySelectorAll(
                'button, input[type="button"], input[type="submit"]'
            )) {{
                const t = (el.textContent || el.value || '').trim().toLowerCase();
                const id = (el.id || '').toLowerCase();
                if (isVisible(el) && kws.some(k => t === k || t.startsWith(k + ' '))
                    && !skipIds.some(s => id.includes(s))) {{
                    el.click();
                    return 'body:' + t + '[' + el.id + ']';
                }}
            }}
            return null;
        }}""")
        if result:
            print(f"    → Popup fechado: [{result}]", flush=True)
            await asyncio.sleep(2)
        else:
            break


async def _screenshot(pg, name: str):
    try:
        SS_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%H%M%S")
        path = SS_DIR / f"{name}_{ts}.png"
        await pg.screenshot(path=str(path))
    except Exception:
        pass


async def _logout(pg):
    """Faz logout do SIAFE clicando 'Sair' — preserva cookie de 30 dias."""
    try:
        # Tenta clicar no link "Sair" no topo da página
        r = await pg.evaluate("""() => {
            for (const el of document.querySelectorAll('a, button')) {
                const t = el.textContent.trim().toLowerCase();
                if (t === 'sair' || t === 'logout' || t === 'exit') {
                    el.click(); return 'sair:' + t;
                }
            }
            return null;
        }""")
        if r:
            print(f"    → Logout via click: {r}")
            await _settle(pg, 4000)
        else:
            # Fallback: navega diretamente para URL de logout
            await pg.goto("https://siafe2.fazenda.rj.gov.br/Siafe/logout",
                          wait_until="domcontentloaded", timeout=15_000)
            await _settle(pg, 3000)
            print("    → Logout via URL direta")
    except Exception as exc:
        print(f"    [w] logout: {exc}")


# ── PASSO 1: Login ─────────────────────────────────────────────────────────────

async def _login(pg, ano: int) -> bool:
    usuario = os.environ.get("SIAFE_USER") or os.environ.get("SIAFE_USUARIO","")
    senha   = os.environ.get("SIAFE_PASS") or os.environ.get("SIAFE_SENHA","")
    if not usuario or not senha:
        print("  [ERRO] SIAFE_USER / SIAFE_PASS não definidos no .env")
        return False

    exercicio_val = EXERCICIOS.get(ano, "1")
    print(f"\n  → Login exercício {ano} (SELECT valor={exercicio_val}) …")
    print(f"    Usuário: {usuario[:4]}*** | Credenciais: {'OK' if usuario and senha else 'AUSENTES'}")

    # Se já logado (sessão de exercício anterior), faz logout primeiro
    # Isso garante que o formulário de login com cbxExercicio vai aparecer
    try:
        if pg.url and "login" not in pg.url.lower() and "about:blank" not in pg.url.lower():
            print("    → Sessão anterior detectada — fazendo logout para trocar exercício")
            await _logout(pg)
    except Exception:
        pass

    # Navega para login — espera rede estabilizar (ADF renderiza via JS)
    await pg.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=TIMEOUT)
    await _settle(pg, 6000)

    title = await pg.title()
    print(f"    Página: '{title[:60]}' | URL: {pg.url[:70]}")

    # Fecha popups de sessão expirada/aviso (allow_confirm=True: aceita continuar/confirmar em overlays)
    await _dismiss_popups(pg, allow_confirm=True)
    await _screenshot(pg, f"01_login_{ano}")

    # Debug: elementos do formulário visíveis
    form_els = await pg.evaluate("""() =>
        [...document.querySelectorAll('input,select')].filter(
            e => e.getBoundingClientRect().width > 0
        ).map(e => e.id || e.name || e.type).filter(Boolean).slice(0,15)
    """)
    print(f"    Form elements visíveis: {form_els}")

    # Debug: todos os <select> da página (opções incluídas)
    sel_info = await pg.evaluate("""() =>
        [...document.querySelectorAll('select')].map(s => ({
            id: s.id,
            visible: s.getBoundingClientRect().width > 0,
            opts: [...s.options].map(o => o.value + ':' + o.text.substring(0,18))
        }))
    """)
    print(f"    Selects: {json.dumps(sel_info, ensure_ascii=False)[:600]}")

    # Preenche usuário
    for field_id, value in [
        ("loginBox:itxUsuario::content",  usuario),
        ("loginBox:itxSenhaAtual::content", senha),
    ]:
        filled = False
        try:
            el = pg.locator(f'[id="{field_id}"]').first
            await el.wait_for(state="visible", timeout=5000)
            await el.click(timeout=4000)
            await el.fill("")
            await pg.keyboard.type(value, delay=40)
            await asyncio.sleep(0.3)
            filled = True
        except Exception as exc:
            print(f"    [w] campo {field_id}: {exc}")
        if not filled:
            await pg.evaluate(
                f"""(v) => {{
                    const el = document.getElementById('{field_id}');
                    if (el) {{ el.value = v; ['input','change','blur'].forEach(
                        e => el.dispatchEvent(new Event(e, {{bubbles:true}})));
                    }}
                }}""",
                value,
            )

    # Seleciona cliente (RJ=0) e exercício — ADF af:selectOneChoice
    # Tenta várias estratégias em cascata
    for sel_id, val in [("cbxCliente", "0"), ("cbxExercicio", exercicio_val)]:
        set_ok = False

        # Estratégia 1: Playwright nativo no elemento ::content (seletor CSS direto)
        for css in [
            f'select[id*="{sel_id}::content"]',
            f'[id*="{sel_id}::content"]',
        ]:
            try:
                loc = pg.locator(css).first
                if await loc.count() > 0:
                    await loc.select_option(val, timeout=3000)
                    print(f"    select {sel_id}: native css={css} → {val} ✔")
                    set_ok = True
                    break
            except Exception:
                pass

        # Estratégia 2: JavaScript com múltiplos padrões
        if not set_ok:
            result = await pg.evaluate(f"""(v) => {{
                // Tenta ID com ::content (padrão ADF)
                let el = document.querySelector('[id*="{sel_id}::content"]');
                // Fallback: select filho do container ADF
                if (!el) {{
                    const c = document.querySelector('[id*="{sel_id}"]');
                    el = c ? (c.tagName==='SELECT' ? c : c.querySelector('select')) : null;
                }}
                // Fallback: qualquer select com nome/id parecido
                if (!el) {{
                    for (const s of document.querySelectorAll('select')) {{
                        if ((s.id||'').toLowerCase().includes('{sel_id.lower()}') ||
                            (s.name||'').toLowerCase().includes('{sel_id.lower()}')) {{
                            el = s; break;
                        }}
                    }}
                }}
                if (el && el.tagName === 'SELECT') {{
                    el.value = v;
                    ['change','blur'].forEach(e => el.dispatchEvent(new Event(e,{{bubbles:true}})));
                    return 'js:' + el.id + '=' + el.value + ' (opts=' + el.options.length + ')';
                }}
                // Diagnóstico: mostra todos os selects
                const all = [...document.querySelectorAll('select')].map(s=>s.id||s.name||'?');
                return 'not_found selects=' + JSON.stringify(all);
            }}""", val)
            print(f"    select {sel_id}: {result}")
            if "js:" in result:
                set_ok = True

        # Estratégia 3: ADF Page API — tenta SEMPRE (não só como fallback)
        # ADF mantém estado interno separado do DOM; setValue() dispara PPR correto
        r3 = await pg.evaluate(f"""(v) => {{
            try {{
                if (typeof AdfPage !== 'undefined') {{
                    const comp = AdfPage.PAGE.findComponentByAbsoluteId('loginBox:{sel_id}');
                    if (comp) {{ comp.setValue(v); return 'adf_api:ok'; }}
                    return 'adf_api:comp_not_found';
                }}
            }} catch(e) {{ return 'adf_api:err=' + e.message; }}
            return 'adf_api:AdfPage_undefined';
        }}""", val)
        print(f"    select {sel_id} ADF API: {r3}")

        # Aguarda PPR (partial page render) do ADF estabilizar
        await _settle(pg, 2000)

    # Aguarda ADF finalizar qualquer PPR pendente após todas as seleções
    await _settle(pg, 3000)
    await _screenshot(pg, f"01b_presubmit_{ano}")

    # Clica botão Confirmar/Login
    btn_clicked = False
    for btn_sel in [
        '[id*="loginBox"][id*="btnConfirmar"]',
        '[id*="btnConfirmar"]',
        'input[type="submit"]',
        'button[type="submit"]',
    ]:
        try:
            loc = pg.locator(btn_sel).first
            cnt = await loc.count()
            if cnt > 0:
                txt = (await loc.text_content() or "").strip()[:20]
                print(f"    Clicando botão: '{txt}' ({btn_sel})", flush=True)
                await loc.click(timeout=5000)
                btn_clicked = True
                break
        except Exception:
            pass
    if not btn_clicked:
        print("    [w] Botão não encontrado — pressionando Enter", flush=True)
        await pg.keyboard.press("Enter")

    # Aguarda popup aparecer (popup ADF pode levar ~1s para animar)
    await asyncio.sleep(1.5)
    await _screenshot(pg, f"01c_postclick_{ano}")

    # Aguarda navegação pós-login — poll a cada 2s por até 35s.
    # SIAFE mostra popup "O Sistema está aberto em outra janela" após clicar Ok
    # O botão "Sim" pode estar dentro do loginBox (ADF aninha dialogs no container da página).
    # Portanto NÃO usamos !el.closest('[id*="loginBox"]') — apenas excluímos
    # os campos de credenciais e o próprio btnConfirmar para não fazer loop.
    _deadline_login = asyncio.get_event_loop().time() + 35
    _poll_n = 0
    while asyncio.get_event_loop().time() < _deadline_login:
        if "login" not in pg.url.lower() and "autenticacao" not in pg.url.lower():
            break
        _poll_n += 1

        # Captura screenshot na 2a iteração para diagnóstico (1a é imediatamente pós-click)
        if _poll_n == 2:
            await _screenshot(pg, f"01d_poll2_{ano}")

        # Tenta dispensar popup "O Sistema está aberto em outra janela" / "já está logado"
        # PRIORIDADE 1: clica explicitamente no botão de confirmação ADF (msgDlg::ok)
        # NUNCA clica em botões com 'cancel' no id — são os botões "Não/Cancelar"
        _pop = await pg.evaluate("""() => {
            // Prioridade 1: botão positivo ADF por ID explícito (evita confusão com cancel)
            const okSel = [
                'docPrincipal::msgDlg::ok',
                'docPrincipal:msgDlg:ok',
            ];
            for (const selId of okSel) {
                const btn = document.getElementById(selId);
                if (btn) {
                    const r = btn.getBoundingClientRect();
                    const visible = (r.width > 0 && r.height > 0) || (btn.offsetWidth > 0 && btn.offsetHeight > 0);
                    if (visible) { btn.click(); return 'popup_clicado:msgDlg_ok id=' + btn.id; }
                }
            }
            // Seletor genérico para outros dialogs ADF (nunca cancel)
            const okGeneric = document.querySelector('[id$="::msgDlg::ok"],[id$=":msgDlg:ok"]');
            if (okGeneric) {
                const r = okGeneric.getBoundingClientRect();
                if ((r.width > 0 && r.height > 0) || (okGeneric.offsetWidth > 0 && okGeneric.offsetHeight > 0)) {
                    okGeneric.click();
                    return 'popup_clicado:msgDlg_ok_generic id=' + okGeneric.id;
                }
            }

            // Prioridade 2: varredura por texto — exclui botões de cancelar/não
            const LABELS = ['continuar','prosseguir','continue','sim','yes'];
            const SKIP   = ['itxusuario','itxsenha','btnconfirmar','cancel','nao','não'];
            for (const el of document.querySelectorAll(
                'button, a[role="button"], input[type="button"], input[type="submit"]'
            )) {
                const t = (el.textContent || el.value || '').trim().toLowerCase();
                const id = (el.id || el.className || '').toLowerCase();
                const r = el.getBoundingClientRect();
                const visible = (r.width > 0 && r.height > 0)
                             || (el.offsetWidth > 0 && el.offsetHeight > 0);
                if (visible
                    && LABELS.some(l => t === l || t.startsWith(l))
                    && !SKIP.some(s => id.includes(s) || t === s)) {
                    el.click();
                    return 'popup_clicado:' + t + ' id=' + el.id;
                }
            }
            // Diagnóstico: lista todos os botões visíveis (para debug nos logs)
            const allBtns = [...document.querySelectorAll(
                'button, a[role="button"], input[type="button"], input[type="submit"]'
            )].filter(el => {
                const r = el.getBoundingClientRect();
                return r.width > 0 || el.offsetWidth > 0;
            }).map(el => (el.textContent || el.value || '').trim().substring(0,20) + '[' + el.id.substring(0,30) + ']');
            return allBtns.length ? 'no_match btns=' + JSON.stringify(allBtns.slice(0,8)) : null;
        }""")
        if _pop and _pop.startswith('popup_clicado:'):
            print(f"    → Popup dispensado: [{_pop}]", flush=True)
            await asyncio.sleep(2)
        elif _pop:
            print(f"    → Poll#{_poll_n}: {_pop}", flush=True)
            # Tenta Enter como fallback (confirma dialog com foco padrão ADF)
            await pg.keyboard.press("Enter")
            await asyncio.sleep(2)
        else:
            # Sem botões visíveis — tenta Enter mesmo assim
            await pg.keyboard.press("Enter")
            await asyncio.sleep(2)

    # Detecta tela MFA (SIAFE exige código por e-mail após credenciais válidas)
    if "login" in pg.url.lower():
        body_txt = await pg.evaluate("() => document.body.innerText")
        if "Autenticação Multifator" in body_txt or "código de autenticação" in body_txt.lower():
            masked = await pg.evaluate("""() => {
                const m = (document.body.innerText || '').match(/enviado para ([^\\n.]+)/);
                return m ? m[1].trim() : 'email cadastrado';
            }""")
            print(f"  → MFA detectado — código enviado para {masked}")

            code = await _aguardar_codigo_mfa(masked, ano=ano)

            if code:
                # Preenche campo de código MFA
                # Exclui: campos de login/senha E inputs ocultos (ADF backing fields)
                excluidos = ["itxUsuario", "itxSenhaAtual"]
                filled = await pg.evaluate(f"""(c) => {{
                    const excl = {json.dumps(excluidos)};
                    const inputs = [...document.querySelectorAll('input:not([type="hidden"])')].filter(el => {{
                        const r = el.getBoundingClientRect();
                        if (!r.width || !r.height) return false;
                        const id = el.id || '';
                        return !excl.some(x => id.includes(x));
                    }});
                    if (!inputs.length) return 'nao_encontrado';
                    const el = inputs[0];
                    el.focus(); el.value = c;
                    ['input','change','blur'].forEach(ev => el.dispatchEvent(new Event(ev,{{bubbles:true}})));
                    return el.id + ':count=' + inputs.length;
                }}""", code)
                print(f"  → Código MFA preenchido: {filled}")

                # Marca "Dispensar código neste dispositivo por 30 dias"
                await pg.evaluate("""() => {
                    const cb = document.querySelector('input[type="checkbox"]');
                    if (cb && !cb.checked) cb.click();
                }""")

                # Clica Ok do formulário MFA
                # Usa o ÚLTIMO botão Ok visível (o primeiro é do loginBox, o último é do MFA)
                clicked = await pg.evaluate("""() => {
                    const okBtns = [...document.querySelectorAll(
                        'button, input[type="submit"], input[type="button"]'
                    )].filter(el => {
                        const t = (el.textContent || el.value || '').trim().toLowerCase();
                        const r = el.getBoundingClientRect();
                        return t === 'ok' && r.width > 0 && r.height > 0;
                    });
                    if (!okBtns.length) return 'nao_encontrado';
                    okBtns[okBtns.length - 1].click();
                    return 'clicked_ok_' + okBtns.length + '_of_' + okBtns.length;
                }""")
                print(f"  → Botão MFA Ok: {clicked}")

                try:
                    await pg.wait_for_url(
                        lambda u: "login" not in u.lower() and "autenticacao" not in u.lower(),
                        timeout=15_000,
                    )
                except Exception:
                    pass
                # Aguarda a página pós-MFA carregar completamente antes de sair do _login
                try:
                    await pg.wait_for_load_state("domcontentloaded", timeout=10_000)
                except Exception:
                    pass
                await _settle(pg, 5000)
                print(f"  → Pós-MFA: {pg.url[:80]}")

    await _dismiss_popups(pg, allow_confirm=False)
    await _screenshot(pg, f"02_pos_login_{ano}")

    url = pg.url.lower()
    ok  = "login" not in url and "autenticacao" not in url
    print(f"  → Login {'✔ OK' if ok else '✖ FALHOU'} — {pg.url[:80]}")

    # Diagnóstico pós-login: captura texto da página e mensagens de erro
    page_text = ""
    error_msgs: list = []
    buttons_after: list = []
    try:
        page_text   = await pg.evaluate("() => document.body.innerText.substring(0,3000)")
        error_msgs  = await pg.evaluate("""() =>
            [...document.querySelectorAll(
                '[id*="msg"], [id*="err"], [id*="alert"], [class*="error"], [class*="alert"], ' +
                '[id*="Msg"], [id*="Error"], .xif, .xi5, .xm4'
            )].map(e => e.textContent.replace(/\\s+/g,' ').trim().substring(0,300))
              .filter(t => t.length > 3).slice(0,15)
        """)
        buttons_after = await pg.evaluate("""() =>
            [...document.querySelectorAll('button, input[type="button"], input[type="submit"]')]
            .filter(e => e.getBoundingClientRect().width > 0)
            .map(e => (e.textContent || e.value || '').trim().substring(0,30))
            .filter(Boolean).slice(0,10)
        """)
    except Exception:
        pass

    print(f"    Page text (200): {page_text[:200].replace(chr(10),' ')}")
    if error_msgs:
        print(f"    Error elements: {error_msgs[:5]}")

    # Salva diagnóstico completo em arquivo para inspecção remota
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        dbg = CACHE_DIR / f"debug_login_{ano}.txt"
        dbg.write_text(
            f"timestamp: {datetime.now().isoformat()}\n"
            f"ano: {ano}\n"
            f"exercicio_val: {exercicio_val}\n"
            f"url_pos_load: {pg.url}\n"
            f"title: {title}\n"
            f"form_els: {json.dumps(form_els)}\n"
            f"sel_info: {json.dumps(sel_info, ensure_ascii=False)}\n"
            f"login_ok: {ok}\n"
            f"url_final: {pg.url}\n"
            f"page_text_after: {page_text[:3000]}\n"
            f"error_msgs: {json.dumps(error_msgs, ensure_ascii=False)}\n"
            f"buttons_after: {json.dumps(buttons_after)}\n",
            encoding="utf-8",
        )
    except Exception as _e:
        print(f"    [w] debug save: {_e}")

    return ok


# ── PASSO 2: Navegação até Ordens Bancárias ────────────────────────────────────

async def _ir_obs(pg) -> bool:
    """
    Navega para Execução Financeira → Ordens Bancárias.

    Estratégia primária: CSS a.xgg "OB Orçamentária" (sempre no DOM — direto).
    Estratégia secundária: a.xgh "Execução Financeira" → a.xgg.
    Estratégia terciária: IDs pt1 (rotina SIAFE-rotina-auditoria.md).
    """
    if "ordembancaria" in pg.url.lower():
        return True

    # Aguarda a página estar completamente carregada antes de interagir
    try:
        await pg.wait_for_load_state("domcontentloaded", timeout=12_000)
    except Exception:
        pass
    await _settle(pg, 4000)

    print("  → Navegando: Execução Financeira > Ordens Bancárias …")

    # Estratégia 1: a.xgg direto (sempre renderizado no DOM do ADF)
    r1 = await pg.evaluate("""() => {
        const LABELS = ['Ordens Bancárias', 'OB Orçamentária', 'OB Orcamentaria'];
        for (const css of ['a.xgg', 'a[class*="xgg"]']) {
            for (const el of document.querySelectorAll(css)) {
                const t = el.textContent.trim();
                if (LABELS.some(l => t === l || t.includes('Ordens Banc')) && !el.className.includes('Disabled')) {
                    el.click(); return 'xgg:' + t;
                }
            }
        }
        return null;
    }""")

    if r1:
        print(f"    ✔ Clique direto a.xgg: {r1}")
        await _settle(pg, 6000)
        await _dismiss_popups(pg)
    else:
        # Estratégia 2: via a.xgh "Execução Financeira"
        r2 = await pg.evaluate("""() => {
            for (const el of document.querySelectorAll('a.xgh, a[class*="xgh"]')) {
                if (el.textContent.trim().includes('Financeira') && !el.className.includes('Disabled')) {
                    el.click(); return 'xgh:' + el.textContent.trim();
                }
            }
            return null;
        }""")
        if r2:
            print(f"    ✔ Clique a.xgh: {r2}")
            await _settle(pg, 5000)

        r3 = await pg.evaluate("""() => {
            const LABELS = ['Ordens Bancárias', 'OB Orçamentária'];
            for (const css of ['a.xgg', 'a[class*="xgg"]']) {
                for (const el of document.querySelectorAll(css)) {
                    if (LABELS.some(l => el.textContent.trim().includes(l.substring(0,10))) && !el.className.includes('Disabled')) {
                        el.click(); return 'xgg2:' + el.textContent.trim();
                    }
                }
            }
            return null;
        }""")
        if r3:
            print(f"    ✔ Clique a.xgg após xgh: {r3}")
            await _settle(pg, 6000)
        else:
            # Estratégia 3: expansão por CSS + busca por texto (robusto a mudanças de índice)
            # 3a: tenta expandir "Execução" / "Execução Financeira" via IDs pt1 conhecidos
            for eid, lbl in [
                ("pt1:pt_np4:1:pt_cni6::disclosureAnchor", "Execução"),
                ("pt1:pt_np3:1:pt_cni4::disclosureAnchor", "Execução Financeira"),
            ]:
                try:
                    loc = pg.locator(f'[id="{eid}"]').first
                    if await loc.count() > 0:
                        await loc.click(timeout=5000, force=True)
                        print(f"    ✔ Clique pt1: {lbl}")
                        await asyncio.sleep(3)
                        await _dismiss_popups(pg)
                    else:
                        print(f"    [w] pt1 não achou: {eid}")
                except Exception as exc:
                    print(f"    [w] pt1 erro {lbl}: {exc}")

            # 3b: busca "Ordens Bancárias" por texto em QUALQUER elemento (índice não hardcoded)
            r_ob = await pg.evaluate("""() => {
                const LABELS = ['Ordens Bancárias', 'OB Orçamentária', 'OB Orcamentaria', 'Ordem Bancária'];
                for (const el of document.querySelectorAll('a, li, span, div')) {
                    const t = el.textContent.trim();
                    const r = el.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0 && LABELS.some(l => t === l || t.startsWith('Ordens Banc'))) {
                        el.click(); return 'text_scan:' + t;
                    }
                }
                // Fallback: procura pt1:pt_np2:N:pt_cni3 varrendo índices 0-25
                for (let i = 0; i <= 25; i++) {
                    const el = document.querySelector('[id="pt1:pt_np2:'+i+':pt_cni3"]');
                    if (el) {
                        const t = el.textContent.trim();
                        if (t.includes('Ordens') || t.includes('OB')) {
                            el.click(); return 'pt1_scan_idx'+i+':'+t;
                        }
                    }
                }
                return null;
            }""")
            if r_ob:
                print(f"    ✔ OB encontrado: {r_ob}")
                await _settle(pg, 5000)
            else:
                print("    [w] pt1 OB não encontrado por texto")

    await _screenshot(pg, "03_tela_obs")

    # Verifica se chegou na tela de OBs
    ok_url  = "ordembancaria" in pg.url.lower() or "ordemBancaria" in pg.url
    ok_tbl  = await pg.evaluate("() => !!document.querySelector('[id*=\"tblOrdemBancaria\"], [id*=\"tblOBOrc\"]')")
    ok_text = await pg.evaluate("() => document.body.innerText.includes('Número') && document.body.innerText.includes('Emissão')")
    ok = ok_url or ok_tbl or ok_text
    print(f"  → Tela OB: {'✔ OK' if ok else '✖ FALHOU'} (url={ok_url}, tbl={ok_tbl})")
    if not ok:
        # Debug: lista items do menu
        items = await pg.evaluate("() => [...document.querySelectorAll('a.xgg, a.xgh, a.xyo')].map(e => e.textContent.trim()).filter(Boolean).slice(0,30)")
        print(f"    Menu items: {items}")
    return ok


# ── Estratégia alternativa: Lista de Favorecido para OB ───────────────────────

async def _ir_lista_favorecido(pg) -> bool:
    """Alternativa: menu 'Lista de Favorecido para OB' — permite filtrar por CNPJ diretamente."""
    print("  → Tentando 'Lista de Favorecido para OB' …")
    r = await pg.evaluate("""() => {
        const LABELS = ['Lista de Favorecido para OB', 'Lista Favorecido'];
        for (const el of document.querySelectorAll('a.xgg, a[class*="xgg"], a')) {
            const t = el.textContent.trim();
            if (LABELS.some(l => t.includes(l.substring(0,15))) && !el.className.includes('Disabled')) {
                el.click(); return t;
            }
        }
        return null;
    }""")
    if r:
        print(f"    ✔ Clique lista favorecido: {r}")
        await _settle(pg, 6000)
        await _dismiss_popups(pg)
        return True
    return False


# ── PASSO 3: Filtrar por CNPJ ─────────────────────────────────────────────────

_JS_APLICAR_FILTRO_CNPJ = """
(cnpj) => {
    // 1. Tenta abrir o acordeão de filtro
    const filterAnchors = [
        document.querySelector('[id*="tblOrdemBancaria"][id*="pnlAccordionDec_afrCl0"]'),
        document.querySelector('[id*="tblOrdemBancaria"][id*="sdtFilter::disAcr"]'),
        document.querySelector('[id*="pnlAccordion"][id*="filter"]'),
    ];
    for (const el of filterAnchors) {
        if (el && el.getBoundingClientRect().width > 0) { el.click(); break; }
    }

    // 2. Espera ~ 500ms e acha campo de favorecido
    return new Promise(resolve => setTimeout(() => {
        const candidatos = [...document.querySelectorAll('input[type="text"], input:not([type])')].filter(el => {
            const id  = (el.id  || '').toLowerCase();
            const ph  = (el.placeholder || '').toLowerCase();
            const lbl = document.querySelector('label[for="'+el.id+'"]');
            const lt  = lbl ? lbl.textContent.toLowerCase() : '';
            return (id.includes('favorecido') || id.includes('cnpj') || id.includes('credor') ||
                    ph.includes('favorecido') || ph.includes('cnpj') || lt.includes('cnpj') ||
                    lt.includes('favorecido')) && el.getBoundingClientRect().width > 0;
        });
        if (!candidatos.length) { resolve(null); return; }
        const el = candidatos[0];
        el.focus();
        el.value = '';
        el.dispatchEvent(new Event('input', {bubbles:true}));
        el.value = cnpj;
        ['input','change','blur','keyup'].forEach(e => el.dispatchEvent(new Event(e, {bubbles:true})));
        resolve(el.id || 'ok');
    }, 600));
}
"""

async def _filtrar_por_cnpj(pg, cnpj: str = None, cnpj_fmt: str = None) -> bool:
    cnpj = cnpj or CNPJ
    cnpj_fmt = cnpj_fmt or CNPJ_FMT
    print(f"  → Filtrando por CNPJ {cnpj_fmt} …", flush=True)

    # Tenta abrir acordeão primeiro com clique nativo
    for acc_id in [
        "pt1:tblOrdemBancaria:pnlAccordionDec_afrCl0",
        "pt1:tblOrdemBancaria:sdtFilter::disAcr",
    ]:
        try:
            loc = pg.locator(f'[id*="{acc_id.split(":")[-1]}"]').first
            if await loc.count() > 0:
                await loc.click(timeout=4000, force=True)
                await asyncio.sleep(2)
                break
        except Exception:
            pass

    # Tenta preencher campo
    for cnpj_val in [cnpj_fmt, cnpj]:
        filled = await pg.evaluate(_JS_APLICAR_FILTRO_CNPJ, cnpj_val)
        if filled:
            print(f"    ✔ Campo preenchido ({cnpj_val}): {filled}")
            break
    else:
        print("    [WARN] Campo CNPJ não encontrado — lerá todos e filtrará em Python")
        return False

    await asyncio.sleep(2)

    # Clica Pesquisar
    await pg.evaluate("""() => {
        for (const el of document.querySelectorAll('button, a, input[type="button"]')) {
            const t = (el.textContent || el.value || '').trim().toLowerCase();
            if ((t.includes('pesquis') || t.includes('filtrar') || t.includes('consult')) &&
                el.getBoundingClientRect().width > 0) { el.click(); return t; }
        }
    }""")

    await _settle(pg, 8000)
    await _dismiss_popups(pg)
    await _screenshot(pg, "04_pos_filtro")
    return True


# ── PASSO 4: Ler todas as páginas da tabela ────────────────────────────────────

_JS_LER_GRADE = r"""
() => {
    const container = (
        document.querySelector('[id*="tblOrdemBancaria"]') ||
        document.querySelector('[id*="tblOBOrc"]') ||
        document.querySelector('table')
    );
    if (!container) return {found: false, header: [], rows: []};

    const tbl  = container.tagName === 'TABLE' ? container : (container.querySelector('table') || container);
    const header = [];
    const rows   = [];

    for (const tr of tbl.querySelectorAll('tr')) {
        const cells = [...tr.querySelectorAll('td,th')].map(c => c.textContent.replace(/\s+/g,' ').trim());
        if (!cells.some(c => c.length > 0)) continue;

        // Detecta cabeçalho (contém "Número" e "Emissão")
        if (!header.length && cells.some(c => /N[uú]mero/i.test(c)) && cells.some(c => /Emiss/i.test(c))) {
            header.push(...cells);
            continue;
        }
        // Detecta linha de dados (número OB começa com 4 dígitos + OB)
        if (/^\d{4}(OB|ob)\d+/.test(cells[0] || '')) {
            rows.push(cells);
        }
    }
    return {found: true, header, rows};
}
"""

_JS_PROXIMA_PAGINA = r"""
() => {
    // Tenta vários seletores para botão próxima página
    const sels = [
        '[id*="tblOrdemBancaria"][id*="next"]',
        '[id*="tblOrdemBancaria"][id*="Next"]',
        'a[title="Próxima Página"]',
        'a[title="Next Page"]',
        'button[title*="Próx"]',
    ];
    for (const s of sels) {
        const el = document.querySelector(s);
        if (el && !el.disabled && !el.className?.includes('Disabled') && el.getBoundingClientRect().width > 0) {
            el.click(); return 'sel:' + s.substring(0, 40);
        }
    }
    // Fallback: texto ">" ou ">>"
    for (const el of document.querySelectorAll('a, button')) {
        const t = el.textContent.trim();
        const ti = (el.title || '').toLowerCase();
        const r  = el.getBoundingClientRect();
        if (r.width > 0 && !el.disabled && (t === '>' || t === '>>' || ti.includes('próx') || ti.includes('next'))) {
            el.click(); return 'txt:' + t;
        }
    }
    return null;
}
"""


async def _ler_tabela(pg) -> tuple[list, list]:
    header: list = []
    all_rows: list = []
    pagina = 1

    while True:
        res = await pg.evaluate(_JS_LER_GRADE)
        if not res.get("found"):
            print(f"    [WARN] Tabela não encontrada na pág {pagina}")
            break

        h    = res.get("header", [])
        rows = res.get("rows", [])

        if h and not header:
            header = h
            print(f"    Colunas: {header}")

        all_rows.extend(rows)
        print(f"    → Pág {pagina}: {len(rows)} linhas de OB")

        if not rows:
            break

        # Tenta avançar página
        nxt = await pg.evaluate(_JS_PROXIMA_PAGINA)
        if not nxt:
            break

        print(f"    → Próxima página: {nxt}")
        pagina += 1
        await _settle(pg, 4000)
        await _dismiss_popups(pg)

    return header, all_rows


# ── Parse ─────────────────────────────────────────────────────────────────────

def _parse_rows(header: list, rows: list, cnpj_filter: Optional[str] = None) -> list[dict]:
    # Mapa de coluna por nome (case-insensitive)
    idx = {h.lower().strip(): i for i, h in enumerate(header)}

    def cell(row: list, *keys: str) -> str:
        for k in keys:
            i = idx.get(k.lower().strip())
            if i is not None and i < len(row):
                v = row[i].strip()
                if v and v != "\xa0":
                    return v
        # Posições fallback (confirmadas pelo SIAFE-rotina-auditoria.md)
        POS = {
            "número":0, "ug emitente":1, "ug pagadora":2,
            "data emissão":3, "data emissao":3,
            "status":4, "tipo":5, "tipo de ob":6,
            "favorecido(cnpj)":7, "nome do favorecido":8,
            "gd":9, "processo":10, "re":11, "pd":12,
            "status de envio":13, "valor":14,
        }
        for k in keys:
            p = POS.get(k.lower().strip())
            if p is not None and p < len(row):
                v = row[p].strip()
                if v and v != "\xa0":
                    return v
        return ""

    records = []
    for row in rows:
        num = cell(row, "número", "numero")
        if not num or not re.match(r"^\d{4}(OB|ob)\d+", num):
            continue

        fav_raw = cell(row, "favorecido(cnpj)", "favorecido", "cnpj")
        fav     = re.sub(r"\D", "", fav_raw)[:14]
        if cnpj_filter and fav != cnpj_filter:
            continue

        data_s = cell(row, "data emissão", "data emissao", "data")
        dt     = _parse_date(data_s)
        valor  = _parse_money(cell(row, "valor", "value"))

        records.append({
            "numero_ob":       num,
            "ug_emitente":     cell(row, "ug emitente", "ug"),
            "ug_pagadora":     cell(row, "ug pagadora"),
            "data_emissao":    data_s,
            "ano":             dt.year  if dt else None,
            "mes":             dt.month if dt else None,
            "favorecido_cnpj": fav,
            "favorecido_nome": cell(row, "nome do favorecido", "nome do credor", "nome"),
            "valor":           valor,
            "processo":        cell(row, "processo", "processo sei"),
            "status":          cell(row, "status"),
            "tipo_ob":         cell(row, "tipo de ob", "tipo ob"),
        })

    return records


# ── Coleta de um exercício (modo UG — todas as empresas) ──────────────────────

async def _coletar_ug_ano(browser, ug_code: str, ano: int) -> list[dict]:
    """
    Coleta TODAS as OBs de uma UG específica para um ano, SEM filtro de CNPJ.
    Usado no modo SIAFE_MODO=todos_ugs para varrer todos os órgãos do SIAFE.
    """
    ug_nome = _ug_nome(ug_code)
    print(f"\n  → UG {ug_code} ({ug_nome}) / Ano {ano}", flush=True)

    ctx = browser.contexts[0] if browser.contexts else await browser.new_context(
        viewport={"width": 1366, "height": 900}, locale="pt-BR",
        timezone_id="America/Sao_Paulo", ignore_https_errors=True,
    )
    page = await ctx.new_page()
    await page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

    try:
        ok = await _login(page, ano)
        if not ok:
            await _screenshot(page, f"ERRO_login_ug{ug_code}_{ano}")
            return []

        nav_ok = await _ir_obs(page)
        if not nav_ok:
            print(f"    [ERRO] Não chegou na tela de OBs para UG {ug_code}/{ano}")
            return []

        # Aplica filtro de UG (não de CNPJ)
        await _aplicar_filtro_ug(page, ug_code)

        # Lê todas as páginas
        header, rows = await _ler_tabela(page)
        print(f"    → Linhas brutas: {len(rows)}")

        # Parse sem filtro CNPJ — captura TODOS os favorecidos
        records = _parse_rows(header, rows, cnpj_filter=None)
        for r in records:
            r.setdefault("ug_coleta", ug_code)
            r.setdefault("ug_coleta_nome", ug_nome)
            r["categoria"] = f"siafe_ug_{ug_code}"

        total = sum(r["valor"] for r in records)
        print(f"    → OBs UG {ug_code}: {len(records)} | {_brl(total)}", flush=True)

        # Salva arquivo por UG × ano
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        out = CACHE_DIR / f"obs_ug{ug_code}_{ano}.json"
        out.write_text(json.dumps({
            "ano": ano, "ug_codigo": ug_code, "ug_nome": ug_nome,
            "coleta": datetime.now().isoformat(),
            "total_obs": len(records), "total_valor": total,
            "header": header, "obs": records,
        }, ensure_ascii=False, indent=2), encoding="utf-8")

        return records

    except Exception as exc:
        print(f"    [ERRO] UG {ug_code}/{ano}: {exc}")
        import traceback; traceback.print_exc()
        return []
    finally:
        try:
            await page.close()
        except Exception:
            pass


# ── Coleta de um exercício (modo empresa/CNPJ) ────────────────────────────────

async def _coletar_exercicio(browser, ano: int, empresa: dict | None = None) -> list[dict]:
    if empresa is None:
        empresa = {"cnpj": CNPJ, "cnpj_fmt": CNPJ_FMT, "nome": NOME_EMP, "categoria": "mgs_clean_real"}
    _cnpj     = empresa["cnpj"]
    _cnpj_fmt = empresa.get("cnpj_fmt", _cnpj)
    _nome     = empresa.get("nome", _cnpj)
    _cat      = empresa.get("categoria", f"obs_{_cnpj}")
    print(f"\n{'━'*56}")
    print(f"  Exercício {ano}")
    print(f"{'━'*56}")

    ctx  = browser.contexts[0] if browser.contexts else await browser.new_context(
        viewport={"width":1366,"height":900},
        locale="pt-BR",
        timezone_id="America/Sao_Paulo",
        ignore_https_errors=True,
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
    )
    page = await ctx.new_page()
    await page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

    try:
        # 1. Login
        ok = await _login(page, ano)
        if not ok:
            await _screenshot(page, f"ERRO_login_{ano}")
            return []

        # 2. Navegar até OBs
        nav_ok = await _ir_obs(page)
        if not nav_ok:
            # Tenta alternativa: Lista de Favorecido para OB
            nav_ok = await _ir_lista_favorecido(page)
            if not nav_ok:
                print(f"  [ERRO] Não chegou na tela de OBs para {ano}")
                await _screenshot(page, f"ERRO_nav_{ano}")
                return []

        # 3. Filtrar pelo CNPJ
        filtrou = await _filtrar_por_cnpj(page, _cnpj, _cnpj_fmt)
        cnpj_f  = None if filtrou else _cnpj

        # 4. Ler tabela (todas as páginas)
        header, rows = await _ler_tabela(page)
        print(f"  → Total linhas lidas: {len(rows)}")

        # 5. Parse
        records = _parse_rows(header, rows, cnpj_filter=cnpj_f)
        for r in records:
            r.setdefault("empresa_cnpj", _cnpj)
            r.setdefault("empresa_nome", _nome)
        total   = sum(r["valor"] for r in records)
        print(f"  → OBs {_nome[:30]}: {len(records)} | {_brl(total)}", flush=True)

        # 6. Salva por ano
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cnpj_safe = re.sub(r'[^\w]', '_', _cnpj)
        out = CACHE_DIR / f"obs_{cnpj_safe}_{ano}.json"
        out.write_text(
            json.dumps({
                "ano": ano, "cnpj": _cnpj, "nome_empresa": _nome,
                "coleta": datetime.now().isoformat(),
                "total_linhas_brutas": len(rows),
                "total_obs_filtradas": len(records),
                "total_valor": total,
                "header": header,
                "obs": records,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"  → Salvo: {out}")
        return records

    except Exception as exc:
        print(f"  [ERRO FATAL] {ano}: {exc}")
        import traceback
        traceback.print_exc()
        await _screenshot(page, f"ERRO_fatal_{ano}")
        return []
    finally:
        try:
            await page.close()
        except Exception:
            pass


# ── Relatório Markdown ─────────────────────────────────────────────────────────

def _gerar_relatorio_md(todas: list[dict], anos: list[int]) -> str:
    total_g = sum(ob["valor"] for ob in todas)
    coleta  = datetime.now().strftime("%Y-%m-%d %H:%M")

    L = [
        f"# ORDENS BANCÁRIAS PAGAS — {NOME_EMP}",
        "",
        f"- **CNPJ:** {CNPJ_FMT}",
        f"- **Fonte:** SIAFE2 — Execução Financeira > Ordens Bancárias",
        f"- **Coleta:** {coleta}",
        f"- **Exercícios cobertos:** {', '.join(str(a) for a in sorted(anos))}",
        f"- **Total pago:** {_brl(total_g)} em {len(todas)} Ordens Bancárias",
        "",
        "> Múltiplas OBs por mês são normais — cada nota fiscal/competência gera uma OB separada.",
        "",
        "---", "",
    ]

    # Seção 1: Resumo por ano
    L += ["## 1. Resumo por Ano", "", "| Ano | OBs | Total Pago |", "|-|--:|--:|"]
    for ano in sorted(anos):
        obs_a = [ob for ob in todas if ob.get("ano") == ano]
        L.append(f"| {ano} | {len(obs_a)} | {_brl(sum(o['valor'] for o in obs_a))} |")
    L += [f"| **TOTAL** | **{len(todas)}** | **{_brl(total_g)}** |", "", "---", ""]

    # Seção 2: Matriz Órgão × Ano
    struct: dict = defaultdict(lambda: defaultdict(list))
    for ob in todas:
        if ob.get("ano") and ob.get("ug_emitente"):
            struct[ob["ano"]][ob["ug_emitente"]].append(ob)

    todos_ugs  = sorted(
        {ob.get("ug_emitente","") for ob in todas if ob.get("ug_emitente")},
        key=lambda u: -sum(ob["valor"] for ob in todas if ob.get("ug_emitente") == u),
    )
    todos_anos = sorted(anos)

    L += ["## 2. Matriz Órgão × Ano", ""]
    hdr = "| Órgão | UG |" + "".join(f" {a} |" for a in todos_anos) + " TOTAL |"
    sep = "|---|---|" + "---:|" * (len(todos_anos) + 1)
    L += [hdr, sep]
    totcol = {a: 0.0 for a in todos_anos}
    for ug in todos_ugs:
        nome = _ug_nome(ug)[:34]
        row  = f"| {nome} | {ug} |"
        tot  = 0.0
        for a in todos_anos:
            v = sum(ob["valor"] for ob in struct[a].get(ug, []))
            totcol[a] += v; tot += v
            row += (f" {_brl(v)} |" if v else " — |")
        row += f" {_brl(tot)} |"
        L.append(row)
    tr = "| **TOTAL** | — |"
    tt = 0.0
    for a in todos_anos:
        tr += f" {_brl(totcol[a])} |"; tt += totcol[a]
    L += [tr + f" {_brl(tt)} |", "", "---", ""]

    # Seção 3: Por órgão → por mês → cada OB individual
    L += [
        "## 3. Detalhamento por Órgão — Cada OB por Mês",
        "",
        "> Cada linha = uma Ordem Bancária individual com seu número único.",
        "",
    ]
    for ano in todos_anos:
        L += [f"### === Exercício {ano} ===", ""]
        ugs_ano = sorted(struct[ano].keys(),
                         key=lambda u: -sum(ob["valor"] for ob in struct[ano][u]))
        for ug in ugs_ano:
            obs_ug  = sorted(struct[ano][ug],
                             key=lambda o: (o.get("mes") or 0, o.get("data_emissao") or ""))
            nome    = _ug_nome(ug)
            total_u = sum(ob["valor"] for ob in obs_ug)
            L += [
                f"#### {nome} (UG {ug}) — {_brl(total_u)} — {len(obs_ug)} OBs",
                "",
                "| Mês | Nº OB | Data | Valor (R$) | Processo SEI | Status |",
                "|---|---|---|---:|---|---|",
            ]
            por_mes: dict = defaultdict(list)
            for ob in obs_ug:
                if ob.get("mes"):
                    por_mes[ob["mes"]].append(ob)

            for mes in sorted(por_mes.keys()):
                obs_m   = sorted(por_mes[mes], key=lambda o: o.get("data_emissao") or "")
                sub     = sum(ob["valor"] for ob in obs_m)
                mes_abr = f"{MESES_PT[mes][:3]}/{ano}"
                for i, ob in enumerate(obs_m):
                    ml = mes_abr if i == 0 else ""
                    p  = (ob.get("processo") or "")[:32]
                    st = (ob.get("status") or "")[:14]
                    v  = f"{ob['valor']:,.2f}".replace(",","X").replace(".",",").replace("X",".")
                    L.append(f"| {ml} | {ob['numero_ob']} | {ob['data_emissao']} | {v} | {p} | {st} |")
                sv = f"{sub:,.2f}".replace(",","X").replace(".",",").replace("X",".")
                L.append(f"| **{mes_abr} sub** | | | **{sv}** | {len(obs_m)} OBs | |")

            tv = f"{total_u:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            L += [f"| **TOTAL {ano}** | | | **{tv}** | {len(obs_ug)} OBs | |", ""]
        L += ["---", ""]

    # Seção 4: Resumo mensal geral
    L += ["## 4. Resumo Mensal — Todos os Órgãos", ""]
    for ano in todos_anos:
        obs_a   = [ob for ob in todas if ob.get("ano") == ano]
        total_a = sum(ob["valor"] for ob in obs_a)
        L += [f"### Ano {ano}", "", "| Mês | OBs | Total Pago | % do Ano |", "|-|--:|--:|--:|"]
        por_mes: dict = defaultdict(list)
        for ob in obs_a:
            if ob.get("mes"):
                por_mes[ob["mes"]].append(ob)
        for mes in sorted(por_mes.keys()):
            obs_m = por_mes[mes]
            v     = sum(ob["valor"] for ob in obs_m)
            pct   = v / total_a * 100 if total_a else 0
            vf    = f"{v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
            L.append(f"| {MESES_PT[mes]}/{ano} | {len(obs_m)} | {vf} | {pct:.1f}% |")
        taf = f"{total_a:,.2f}".replace(",","X").replace(".",",").replace("X",".")
        L += [f"| **TOTAL** | **{len(obs_a)}** | **{taf}** | 100% |", "", ""]

    return "\n".join(L)


# ── Excel navegável ───────────────────────────────────────────────────────────

def _salvar_excel(obs: list[dict]):
    """Gera Excel com aba geral + uma aba por ano, formatação básica."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        f_xl = CACHE_DIR / "mgsclean_obs_todas.xlsx"
        wb = openpyxl.Workbook()

        HDR = ["Ano", "Mês", "Número OB", "Data Emissão", "UG Código", "UG Nome",
               "Favorecido CNPJ", "Favorecido Nome", "Valor (R$)", "Tipo OB",
               "Status", "Processo"]
        HDR_FILL = PatternFill("solid", fgColor="2F5597")
        HDR_FONT = Font(bold=True, color="FFFFFF")

        def _fill_sheet(ws, rows):
            ws.append(HDR)
            for cell in ws[1]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="center")
            for ob in rows:
                dt = ob.get("data_emissao","")
                ws.append([
                    ob.get("ano"), ob.get("mes"),
                    ob.get("numero_ob"), dt,
                    ob.get("ug_emitente"), _ug_nome(ob.get("ug_emitente","")),
                    ob.get("favorecido_cnpj"), ob.get("favorecido_nome"),
                    ob.get("valor", 0.0),
                    ob.get("tipo_ob"), ob.get("status"), ob.get("processo"),
                ])
            # Formata coluna de valor como moeda
            val_col = 9
            for row in ws.iter_rows(min_row=2, min_col=val_col, max_col=val_col):
                for cell in row:
                    cell.number_format = '#,##0.00'
            # Auto-width (approximate)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
            # Freeze header row
            ws.freeze_panes = "A2"

        # Aba geral
        ws_all = wb.active
        ws_all.title = "Todas OBs"
        _fill_sheet(ws_all, sorted(obs, key=lambda o: (o.get("ano",0), o.get("data_emissao",""))))

        # Uma aba por ano
        for ano in sorted({o.get("ano") for o in obs if o.get("ano")}, reverse=True):
            ws_ano = wb.create_sheet(str(ano))
            _fill_sheet(ws_ano, [o for o in obs if o.get("ano") == ano])

        wb.save(str(f_xl))
        print(f"✔ Excel: {f_xl} ({len(obs)} OBs)")

    except Exception as exc:
        print(f"  [w] Excel falhou: {exc}")


# ── Persistência no compliance.db ────────────────────────────────────────────

def _salvar_no_db(obs: list[dict]):
    """
    Salva OBs reais no compliance.db.
    Remove estimativas ('mgs_clean_auditoria') dos exercícios coletados antes de inserir.
    """
    try:
        import sqlite3
        db_path = REPO_ROOT / "data" / "compliance.db"
        if not db_path.exists():
            print("  [w] compliance.db não encontrado — pulando inserção DB")
            return

        conn = sqlite3.connect(str(db_path))
        cur  = conn.cursor()
        now  = datetime.now().isoformat()

        anos_reais = sorted({ob["ano"] for ob in obs if ob.get("ano")})

        # Remove estimativas dos exercícios que vamos inserir com dados reais
        for ano in anos_reais:
            cur.execute(
                "DELETE FROM ordens_bancarias WHERE exercicio=? AND categoria='mgs_clean_auditoria'",
                (ano,)
            )
            print(f"    → Estimativas removidas para {ano}: {cur.rowcount} registros")

        # Remove coletas anteriores do mesmo CNPJ para este ano
        for ano in anos_reais:
            cnpjs_reais = {ob.get("empresa_cnpj") for ob in obs if ob.get("ano") == ano and ob.get("empresa_cnpj")}
            for cnpj_r in cnpjs_reais:
                cur.execute(
                    "DELETE FROM ordens_bancarias WHERE exercicio=? AND favorecido_cpf=? AND categoria NOT LIKE '%auditoria%'",
                    (ano, cnpj_r)
                )

        # Insere OBs reais
        for ob in obs:
            cur.execute("""
                INSERT INTO ordens_bancarias
                    (numero_ob, data_emissao, ug_codigo, ug_nome, favorecido_cpf,
                     favorecido_nome, valor, tipo_ob, status, numero_processo,
                     exercicio, coletado_em, created_at, updated_at, categoria, raw_json)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                ob.get("numero_ob"),
                ob.get("data_emissao"),
                ob.get("ug_emitente"),
                _ug_nome(ob.get("ug_emitente","")) if ob.get("ug_emitente") else None,
                ob.get("favorecido_cnpj"),
                ob.get("favorecido_nome"),
                ob.get("valor", 0.0),
                ob.get("tipo_ob"),
                ob.get("status"),
                ob.get("processo"),
                ob.get("ano"),
                now, now, now,
                ob.get("categoria", ob.get("empresa_categoria", "obs_siafe")),
                json.dumps(ob, ensure_ascii=False),
            ))

        conn.commit()
        conn.close()
        print(f"  ✔ DB: {len(obs)} OBs reais inseridas ({anos_reais})")

    except Exception as exc:
        print(f"  [w] Erro DB: {exc}")


# ── Índice de progresso da coleta ────────────────────────────────────────────

_PROGRESS_FILE = REPO_ROOT / "data" / "sei_cache" / "obs_progress.json"

def _carregar_progresso() -> dict:
    """Carrega obs_progress.json para saber quais pares (empresa, ano) já foram coletados."""
    if _PROGRESS_FILE.exists():
        try:
            return json.loads(_PROGRESS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"empresas": {}, "historico_runs": []}

def _ja_coletado(progresso: dict, cnpj: str, ano: int) -> bool:
    """Retorna True se o par (cnpj, ano) já foi coletado com sucesso."""
    _skip = os.environ.get("SIAFE_SKIP_COLETADOS", "true").lower() != "false"
    if not _skip:
        return False
    emp = progresso.get("empresas", {}).get(cnpj, {})
    status = emp.get("anos", {}).get(str(ano), {}).get("status", "pendente")
    return status == "coletado"

def _atualizar_progresso(progresso: dict, cnpj: str, nome: str, ano: int,
                          total_obs: int, total_valor: float, run_id: str = ""):
    """Marca um par (cnpj, ano) como coletado no índice de progresso."""
    now = datetime.now().isoformat()
    emp = progresso.setdefault("empresas", {}).setdefault(cnpj, {
        "nome": nome, "anos": {}
    })
    emp["anos"][str(ano)] = {
        "status": "coletado" if total_obs > 0 else "vazio",
        "total_obs": total_obs,
        "total_valor": total_valor,
        "coletado_em": now,
        "run_id": run_id,
    }
    # Atualiza resumo
    todos = [v for e in progresso.get("empresas", {}).values()
             for v in e.get("anos", {}).values()]
    coletados = sum(1 for v in todos if v.get("status") in ("coletado", "vazio"))
    progresso["ultima_atualizacao"] = now
    progresso["resumo"] = {
        "total_empresas": len(progresso.get("empresas", {})),
        "total_pares_empresa_ano": len(todos),
        "coletados": coletados,
        "pendentes": len(todos) - coletados,
        "percentual_completo": round(coletados / len(todos) * 100, 1) if todos else 0.0,
    }

def _salvar_progresso(progresso: dict):
    """Persiste o índice de progresso em disco."""
    try:
        _PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
        _PROGRESS_FILE.write_text(
            json.dumps(progresso, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as exc:
        print(f"  [w] Progresso não salvo: {exc}")

def _imprimir_progresso(progresso: dict):
    """Imprime resumo do progresso de coleta."""
    r = progresso.get("resumo", {})
    print(f"\n  ► Progresso coleta: {r.get('coletados',0)}/{r.get('total_pares_empresa_ano',0)} pares "
          f"({r.get('percentual_completo',0)}%) — {r.get('pendentes',0)} pendentes")
    for cnpj, emp in progresso.get("empresas", {}).items():
        anos_ok = [a for a, v in emp.get("anos", {}).items() if v.get("status") == "coletado"]
        anos_pend = [a for a, v in emp.get("anos", {}).items() if v.get("status") == "pendente"]
        if anos_ok or anos_pend:
            print(f"    {emp.get('nome','')[:35]} | ✔ {','.join(anos_ok)} | ○ {','.join(anos_pend)}")


# ── Git push automático ────────────────────────────────────────────────────────

def _git_push(arquivos: list[str], mensagem: str):
    try:
        repo = REPO_ROOT
        subprocess.run(["git", "add"] + arquivos, cwd=repo, check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", mensagem], cwd=repo, check=True, capture_output=True)
        r = subprocess.run(
            ["git", "push", "-u", "origin", "HEAD"],
            cwd=repo, capture_output=True, text=True, timeout=60
        )
        if r.returncode == 0:
            print(f"  ✔ Git push OK")
        else:
            print(f"  [w] Git push saída: {r.stderr[:200]}")
    except subprocess.CalledProcessError as exc:
        print(f"  [w] Git erro: {exc}")
    except Exception as exc:
        print(f"  [w] Git exceção: {exc}")


# ── Main ───────────────────────────────────────────────────────────────────────

async def main():
    from playwright.async_api import async_playwright

    SIAFE_USER = os.environ.get("SIAFE_USER") or os.environ.get("SIAFE_USUARIO","")
    SIAFE_PASS = os.environ.get("SIAFE_PASS") or os.environ.get("SIAFE_SENHA","")
    if not SIAFE_USER or not SIAFE_PASS:
        print("\n[ERRO] Credenciais não encontradas!")
        print("Defina SIAFE_USER e SIAFE_PASS no arquivo .env")
        sys.exit(1)

    _cur_year = datetime.now().year
    _anos_env = os.environ.get("SIAFE_ANOS", "")
    _anos_raw = sys.argv[1:] if sys.argv[1:] else _anos_env.split(",")
    anos = sorted(
        [int(a.strip()) for a in _anos_raw if a.strip().isdigit() and int(a.strip()) in EXERCICIOS]
        or [y for y in EXERCICIOS if y <= _cur_year],
        reverse=True,
    )

    # Modo de coleta: "cnpjs" (padrão) ou "todos_ugs" (varre todos os órgãos SIAFE)
    _modo = os.environ.get("SIAFE_MODO", "cnpjs").strip().lower()
    _ugs_env = os.environ.get("SIAFE_UGS", "").strip()

    empresas = _load_empresas()
    progresso = _carregar_progresso()
    _run_id = os.environ.get("GITHUB_RUN_ID", "local")

    print(f"  Modo: {_modo}")
    print(f"  Empresas: {len(empresas)} ({', '.join(e.get('nome','?')[:20] for e in empresas[:3])}{'...' if len(empresas) > 3 else ''})")
    _imprimir_progresso(progresso)

    print(f"\n{'━'*56}")
    print(f"  JFN — Coleta de Ordens Bancárias SIAFE")
    print(f"  Anos: {anos} | Modo: {_modo}")
    print(f"  Usuário: {SIAFE_USER[:4]}***")
    print(f"{'━'*56}")

    _telegram(f"🟡 Iniciando coleta SIAFE OBs — Modo: {_modo} | Anos: {anos}")

    p = await async_playwright().start()
    browser = None

    # Tenta CDP (Chrome já aberto) primeiro; senão lança Chromium
    try:
        browser = await p.chromium.connect_over_cdp("http://127.0.0.1:9222", timeout=4000)
        print("✔ Chrome CDP (porta 9222)")
    except Exception:
        print("ℹ Lançando Chromium …")
        _ci = bool(os.environ.get("CI") or os.environ.get("GITHUB_ACTIONS"))
        _headless = _ci or os.environ.get("HEADLESS", "false").lower() == "true"
        browser = await p.chromium.launch(
            headless=_headless,
            slow_mo=0 if _headless else 150,
            args=[
                "--no-sandbox", "--disable-dev-shm-usage",
                "--disable-gpu" if _headless else "--start-maximized",
                "--disable-blink-features=AutomationControlled",
            ],
        )
        print("✔ Chromium lançado")

    todas_obs: list[dict] = []
    anos_coletados: list[int] = []

    try:
        todas_por_empresa: dict[str, list] = {}

        if _modo == "todos_ugs":
            # ── Modo varredura total: login, descobre UGs, coleta tudo ──────────────
            # Determina lista de UGs (env SIAFE_UGS, ou descoberta dinâmica, ou estática)
            if _ugs_env:
                ugs_alvo = [u.strip() for u in _ugs_env.split(",") if u.strip()]
                print(f"  UGs via SIAFE_UGS: {ugs_alvo}")
            else:
                # Faz login temporário para descobrir UGs da interface
                _ctx0 = await browser.new_context(viewport={"width":1366,"height":900}, locale="pt-BR",
                    timezone_id="America/Sao_Paulo", ignore_https_errors=True)
                _p0 = await _ctx0.new_page()
                try:
                    if await _login(_p0, anos[0]):
                        if await _ir_obs(_p0):
                            ugs_alvo = await _descobrir_ugs_siafe(_p0)
                        else:
                            ugs_alvo = list(_UG_NOMES.keys())
                    else:
                        ugs_alvo = list(_UG_NOMES.keys())
                finally:
                    await _p0.close()

            print(f"\n  ► Modo todos_ugs: {len(ugs_alvo)} UGs × {len(anos)} anos = {len(ugs_alvo)*len(anos)} sessões")
            _telegram(f"🔄 todos_ugs: {len(ugs_alvo)} UGs × {len(anos)} anos")

            for ano in anos:
                for ug_code in ugs_alvo:
                    obs_ug = await _coletar_ug_ano(browser, ug_code, ano)
                    if obs_ug:
                        todas_obs.extend(obs_ug)
                        if ano not in anos_coletados:
                            anos_coletados.append(ano)
                        # Agrupa por CNPJ favorecido
                        for ob in obs_ug:
                            cnpj_fav = ob.get("favorecido_cnpj", "")
                            if cnpj_fav:
                                todas_por_empresa.setdefault(cnpj_fav, []).append(ob)
                        # Marca UG como coletada no progresso
                        _atualizar_progresso(progresso, f"ug_{ug_code}", _ug_nome(ug_code),
                                             ano, len(obs_ug), sum(o["valor"] for o in obs_ug), _run_id)
                        _salvar_progresso(progresso)
                    await asyncio.sleep(2)

        else:
            # ── Modo padrão: por empresa/CNPJ ────────────────────────────────────
            for ano in anos:
                for empresa in empresas:
                    _cnpj = empresa["cnpj"]
                    _nome = empresa.get("nome", _cnpj)
                    if _ja_coletado(progresso, _cnpj, ano):
                        prev = progresso["empresas"][_cnpj]["anos"][str(ano)]
                        print(f"  ↷ Pulando {_nome[:30]} / {ano} — já coletado ({prev['total_obs']} OBs)")
                        # Recarrega as OBs do arquivo salvo
                        cnpj_safe = re.sub(r'[^\w]', '_', _cnpj)
                        f_prev = CACHE_DIR / f"obs_{cnpj_safe}_{ano}.json"
                        if f_prev.exists():
                            try:
                                prev_data = json.loads(f_prev.read_text(encoding="utf-8"))
                                obs_prev = prev_data.get("obs", [])
                                todas_por_empresa.setdefault(_cnpj, []).extend(obs_prev)
                                todas_obs.extend(obs_prev)
                                if ano not in anos_coletados:
                                    anos_coletados.append(ano)
                            except Exception:
                                pass
                        continue

                    obs_ano = await _coletar_exercicio(browser, ano, empresa)
                    _atualizar_progresso(progresso, _cnpj, _nome, ano,
                                         len(obs_ano), sum(o["valor"] for o in obs_ano), _run_id)
                    _salvar_progresso(progresso)

                    if obs_ano:
                        todas_por_empresa.setdefault(_cnpj, []).extend(obs_ano)
                        todas_obs.extend(obs_ano)
                        if ano not in anos_coletados:
                            anos_coletados.append(ano)
                    await asyncio.sleep(3)

        if not todas_obs:
            msg = "[ERRO] Nenhuma OB coletada — login falhou (MFA expirado/inválido?) ou CNPJ sem dados."
            print(f"\n{msg}")
            print("  → Verifique data/sei_cache/debug_login_*.txt para diagnóstico.")
            _telegram(f"❌ {msg}")
            sys.exit(1)

        # Salva consolidado
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        f_json = CACHE_DIR / "obs_todas.json"
        f_json.write_text(
            json.dumps({
                "coleta": datetime.now().isoformat(),
                "empresas": len(todas_por_empresa),
                "cnpjs_coletados": list(todas_por_empresa.keys()),
                "anos_cobertos": sorted(anos_coletados),
                "total_obs": len(todas_obs),
                "total_valor": sum(ob["valor"] for ob in todas_obs),
                "obs": todas_obs,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"\n✔ Consolidado: {f_json} ({len(todas_obs)} OBs)")

        # Salva por-empresa consolidado
        for cnpj_e, obs_e in todas_por_empresa.items():
            cnpj_safe = re.sub(r'[^\w]', '_', cnpj_e)
            fp = CACHE_DIR / f"obs_{cnpj_safe}_todas.json"
            fp.write_text(json.dumps({
                "cnpj": cnpj_e,
                "coleta": datetime.now().isoformat(),
                "anos_cobertos": sorted(anos_coletados),
                "total_obs": len(obs_e),
                "total_valor": sum(o["valor"] for o in obs_e),
                "obs": obs_e,
            }, ensure_ascii=False, indent=2), encoding="utf-8")

        # Backward compat: arquivo MGS CLEAN legado
        mgs_cnpj = "19088605000104"
        if mgs_cnpj in todas_por_empresa:
            legacy = CACHE_DIR / "mgsclean_obs_todas.json"
            mgs_obs = todas_por_empresa[mgs_cnpj]
            legacy.write_text(json.dumps({
                "cnpj": mgs_cnpj, "nome_empresa": "MGS CLEAN SOLUCOES E SERVICOS LTDA",
                "coleta": datetime.now().isoformat(), "anos_cobertos": sorted(anos_coletados),
                "total_obs": len(mgs_obs), "total_valor": sum(o["valor"] for o in mgs_obs),
                "obs": mgs_obs,
            }, ensure_ascii=False, indent=2), encoding="utf-8")

        # Relatório Markdown
        md = _gerar_relatorio_md(todas_obs, anos_coletados)
        f_md = CACHE_DIR / "mgsclean_obs_resumo.md"
        f_md.write_text(md, encoding="utf-8")
        print(f"✔ Relatório MD: {f_md}")

        total = sum(ob["valor"] for ob in todas_obs)

        print(f"\n{'━'*56}")
        print(f"  TOTAL GERAL PAGO: {_brl(total)}")
        print(f"  OBs coletadas:    {len(todas_obs)}")
        print(f"  Anos:             {', '.join(str(a) for a in sorted(anos_coletados))}")
        print(f"{'━'*56}")

        # Excel navegável
        _salvar_excel(todas_obs)

        # Salva no compliance.db
        print("\n→ Salvando no banco de dados …")
        _salvar_no_db(todas_obs)

        # Git push automático
        print("→ Fazendo git push …")
        _imprimir_progresso(progresso)
        arquivos_json = [str(f_json), str(f_md), str(_PROGRESS_FILE)]
        for cnpj_e in todas_por_empresa:
            cnpj_safe = re.sub(r'[^\w]', '_', cnpj_e)
            arquivos_json += [
                str(CACHE_DIR / f"obs_{cnpj_safe}_todas.json"),
            ] + [str(CACHE_DIR / f"obs_{cnpj_safe}_{a}.json") for a in anos_coletados]
        if (CACHE_DIR / "mgsclean_obs_todas.json").exists():
            arquivos_json.append(str(CACHE_DIR / "mgsclean_obs_todas.json"))
        _git_push(
            arquivos_json,
            f"dados: OBs SIAFE {len(todas_por_empresa)} empresas {'/'.join(str(a) for a in sorted(anos_coletados))} — {len(todas_obs)} OBs {_brl(total)}"
        )

        # Notificação Telegram
        msg = (
            f"✅ Coleta SIAFE concluída!\n"
            f"CNPJ: {CNPJ_FMT}\n"
            f"Anos: {', '.join(str(a) for a in sorted(anos_coletados))}\n"
            f"OBs: {len(todas_obs)}\n"
            f"Total pago: {_brl(total)}\n"
            f"Dados salvos em data/sei_cache/"
        )
        _telegram(msg)
        print("\nPróximo passo:")
        print("  python _SANDBOX/gerar_relatorio_obs_pdf.py")

    finally:
        try:
            await browser.close()
        except Exception:
            pass
        await p.stop()


if __name__ == "__main__":
    asyncio.run(main())
